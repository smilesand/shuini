from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FONT_NAME = "Microsoft YaHei"
THIN_SIDE = Side(style="thin", color="D9DEE8")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTITLE_FILL = PatternFill("solid", fgColor="DCE6F1")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
ALT_FILL = PatternFill("solid", fgColor="F7FBFF")

STATUS_FILLS = {
    "已完成": PatternFill("solid", fgColor="E2F0D9"),
    "部分完成": PatternFill("solid", fgColor="FFF2CC"),
    "未完成/占位": PatternFill("solid", fgColor="F4CCCC"),
    "保留/未接前端": PatternFill("solid", fgColor="D9EAD3"),
}

SETTLEMENT_FILLS = {
    "是": PatternFill("solid", fgColor="E2F0D9"),
    "审慎": PatternFill("solid", fgColor="FFF2CC"),
    "否": PatternFill("solid", fgColor="F4CCCC"),
    "不单独计费": PatternFill("solid", fgColor="F4CCCC"),
    "建议纳入": PatternFill("solid", fgColor="E2F0D9"),
    "暂不建议": PatternFill("solid", fgColor="F4CCCC"),
}


FEATURE_HEADERS = [
    "序号",
    "一级模块",
    "二级模块",
    "功能名称",
    "功能说明",
    "使用角色",
    "入口页面/路由",
    "主要接口/能力",
    "交付状态",
    "建议纳入本次结算",
    "备注",
]

FEATURE_ROWS = [
    [1, "认证与权限", "登录", "用户登录", "输入用户名和密码登录系统，写入本地登录态并获得服务端 session。", "全员", "/login", "POST /api/auth/login", "已完成", "是", "登录成功后进入系统首页"],
    [2, "认证与权限", "退出登录", "安全退出", "退出当前账号，并同步注销服务端 session。", "全员", "全局头部", "POST /api/auth/logout", "已完成", "是", "支持任意页面退出"],
    [3, "认证与权限", "首登提醒", "默认密码提醒", "首次登录后通过 must_reset 状态在首页提示尽快修改默认密码。", "全员", "首页 / 个人中心", "GET /api/auth/profile", "已完成", "是", "后端已保留首次改密状态位"],
    [4, "个人中心", "资料维护", "邮箱/手机号维护", "查看并更新当前用户的邮箱、手机号。", "全员", "/profile", "GET /api/auth/profile; PUT /api/auth/profile", "已完成", "是", "个人信息更新后即时生效"],
    [5, "个人中心", "密码管理", "修改密码", "校验原密码后修改新密码，修改完成后要求重新登录。", "全员", "/profile", "POST /api/auth/change-password", "已完成", "是", "提升账号安全性"],
    [6, "用户管理", "用户列表", "查询用户", "管理员查看用户列表、角色、联系方式、密码状态。", "管理员", "/settings/users", "GET /api/auth/users", "已完成", "是", "支持关键字过滤"],
    [7, "用户管理", "新增用户", "创建账号", "管理员新增普通用户，默认密码为 123456。", "管理员", "/settings/users", "POST /api/auth/users", "已完成", "是", "新用户首次登录需修改密码"],
    [8, "用户管理", "密码重置", "重置用户密码", "管理员为指定用户重置密码。", "管理员", "/settings/users", "POST /api/auth/users/{username}/reset-password", "已完成", "是", "重置后用户需重新设置密码"],
    [9, "用户管理", "删除用户", "删除普通用户", "管理员可删除普通用户账号，管理员账号不可直接删除。", "管理员", "/settings/users", "DELETE /api/auth/users/{username}", "已完成", "是", "避免误删管理员"],
    [10, "项目管理", "项目列表", "查询与搜索项目", "按项目编号或项目名称搜索，并分页展示项目列表和配比数量。", "全员/管理员", "/projects", "GET /api/projects", "已完成", "是", "普通用户默认只看自己创建的项目"],
    [11, "项目管理", "项目新建", "创建项目", "录入项目编号、项目名称和项目要求，生成项目档案。", "全员/管理员", "/projects", "POST /api/projects", "已完成", "是", "项目编号唯一"],
    [12, "项目管理", "项目详情", "查看项目详情", "查看项目基础信息、创建人、创建时间和项目下关联记录。", "全员/管理员", "/projects/:id", "GET /api/projects/{project_id}", "已完成", "是", "项目要求可完整展示"],
    [13, "项目管理", "项目维护", "编辑项目", "修改项目编号、名称和项目要求。", "全员/管理员", "/projects/:id", "PUT /api/projects/{project_id}", "已完成", "是", "编辑后重新加载详情"],
    [14, "项目管理", "项目删除", "删除项目", "删除项目本身，但不级联删除历史配比记录，仅解除项目关联。", "全员/管理员", "/projects; /projects/:id", "DELETE /api/projects/{project_id}", "已完成", "是", "便于保留历史数据"],
    [15, "项目管理", "项目记录", "查看项目内配比记录", "在项目详情页查看项目下全部配比记录，并支持载入和删除。", "全员/管理员", "/projects/:id", "GET /api/projects/{project_id}/records", "已完成", "是", "项目与记录形成对应关系"],
    [16, "项目管理", "快捷新建", "从项目创建 HPC 配比", "在项目详情页一键进入高性能混凝土配比计算，并自动绑定当前项目。", "全员/管理员", "/projects/:id -> /calc/hpc", "前端路由 + projects/records API", "已完成", "是", "便于按项目归档计算结果"],
    [17, "项目管理", "快捷新建", "从项目创建 HPC 试配", "在项目详情页一键进入高性能试配页，并自动绑定当前项目。", "全员/管理员", "/projects/:id -> /adapt/hpc", "前端路由 + POST /api/hpc-trial", "已完成", "是", "与配比设计结果联动"],
    [18, "记录管理", "记录保存", "保存/更新配比记录", "在最终计算步骤保存新记录，或覆盖更新当前已载入记录。", "全员/管理员", "/calc/hpc", "POST /api/records", "已完成", "是", "记录包含项目、类别和材料结果"],
    [19, "记录管理", "历史记录", "搜索/分页/载入/删除", "在右侧汇总面板查看当前类别历史记录，支持关键字搜索、分页、双击载入和删除。", "全员/管理员", "/calc/hpc 右侧侧栏", "GET /api/records; DELETE /api/records/{record_id}", "已完成", "是", "支持按项目过滤"],
    [20, "HPC 配比计算", "水胶比", "自动计算 fcu0 和 W/B", "根据强度等级、胶材 28d 强度和回归系数自动计算配制强度与水胶比。", "全员", "/calc/hpc -> 水胶比", "POST /api/water-binder", "已完成", "是", "输入变化后自动防抖重算"],
    [21, "HPC 配比计算", "回归系数", "手工维护回归系数", "支持直接录入 αa、αb、αc 参与水胶比计算。", "全员", "/calc/hpc -> 水胶比", "POST /api/water-binder", "已完成", "是", "系统内置默认系数"],
    [22, "HPC 配比计算", "回归拟合", "Excel/CSV 拟合系数", "上传 CSV/XLS/XLSX 文件，自动解析并拟合回归系数后回填页面。", "全员", "/calc/hpc -> 水胶比", "POST /api/upload-fit; POST /api/fit-coefficients", "已完成", "是", "适合按试验数据快速拟合"],
    [23, "HPC 配比计算", "砂率选取", "查表选取砂率", "提供不同强度等级与粒径的砂率参考表，支持交叉高亮和输入确认。", "全员", "/calc/hpc -> 砂率选取", "前端页面逻辑", "已完成", "是", "已自动传递到骨料计算"],
    [24, "HPC 配比计算", "骨料用量", "计算粗细骨料", "根据 Vg、骨料密度和砂率计算粗骨料与细骨料用量。", "全员", "/calc/hpc -> 骨料用量", "POST /api/aggregate", "已完成", "是", "支持 Vg 参考范围提示"],
    [25, "HPC 配比计算", "胶凝材料", "计算胶材总量与分项", "支持粉煤灰、矿粉、微珠、硅灰四类掺合料，计算胶材密度、浆体体积和各材料用量。", "全员", "/calc/hpc -> 胶凝材料", "POST /api/binder", "已完成", "是", "能输出水泥与各掺合料具体质量"],
    [26, "HPC 配比计算", "水和外加剂", "计算用水量与外加剂", "根据胶凝总量、水胶比和外加剂掺量计算 m_w、m_a，并汇总每方总量。", "全员", "/calc/hpc -> 水和外加剂", "POST /api/water-admixture", "已完成", "是", "计算完成后可直接保存记录"],
    [27, "HPC 配比计算", "汇总展示", "右侧汇总面板", "实时展示 W/B、砂率、骨料、胶材、水和外加剂合计，并展示错误信息。", "全员", "/calc/hpc 右侧侧栏", "前端页面逻辑", "已完成", "是", "便于快速核对关键结果"],
    [28, "HPC 配比计算", "项目模式", "先选项目后计算", "计算页进入后需先选择项目，并可查看、切换当前项目下已有 HPC 配比记录。", "全员", "/calc/hpc", "GET /api/projects; GET /api/projects/{project_id}/records", "已完成", "是", "确保记录按项目归档"],
    [29, "HPC 试配调整", "工作性实验", "调整胶材/砂率/外加剂", "保持 W/B 不变，按增减量重算每方配合比，并换算指定试拌体积用量。", "全员", "/adapt/hpc -> 工作性实验", "POST /api/hpc-trial", "已完成", "是", "支持坍落度录入和备注说明"],
    [30, "HPC 试配调整", "强度实验", "三组试验与回归分析", "自动生成基准/+Δ/-Δ 三组强度试验配合比，录入 28d 强度后输出回归结果、R² 和推荐 W/B。", "全员", "/adapt/hpc -> 强度实验", "POST /api/hpc-trial", "已完成", "是", "附带小型关系图"],
    [31, "HPC 试配调整", "配合比校正", "参数调整与密度校正", "调整 W/B、胶材用量、砂率、外加剂掺量，输入表观密度后自动生成实验室配合比。", "全员", "/adapt/hpc -> 配合比校正与确认", "POST /api/hpc-trial", "已完成", "是", "支持载入强度实验推荐水胶比"],
    [32, "首页与说明", "首页", "产品首页与入口", "展示系统定位、核心功能卡片，并提供快速开始计算入口。", "全员", "/", "前端页面逻辑", "已完成", "是", "首页会显示默认密码提醒"],
    [33, "首页与说明", "关于页", "计算流程说明", "展示技术栈和 5 步配比计算流程说明。", "全员", "/about", "前端页面逻辑", "已完成", "是", "适合作为系统介绍页"],
    [34, "系统支撑", "菜单权限", "路由与菜单权限控制", "未登录不可访问业务页面；管理员菜单仅对管理员显示。", "全员/管理员", "全局路由与布局", "前端路由守卫", "已完成", "是", "减少越权访问"],
    [35, "系统支撑", "数据权限", "项目/记录隔离", "普通用户仅能查看和操作自己创建的项目和记录，管理员可查看全部数据。", "全员/管理员", "项目、记录、用户模块", "projects/records/database 相关接口", "已完成", "是", "服务端已做权限校验"],
    [36, "系统支撑", "会话管理", "服务端 Session 管理", "登录后创建服务端 session，退出登录时失效，并支持过期清理和密码哈希升级。", "全员/管理员", "登录/退出", "auth/database 相关接口", "已完成", "是", "提升安全性和可维护性"],
    [37, "UHPC 模块", "配比计算", "UHPC 计算页占位", "当前页面仅显示“开发中”提示，尚未接入实际公式、保存和历史能力。", "全员", "/calc/uhpc", "无", "未完成/占位", "否", "当前不建议按完整功能结算"],
    [38, "UHPC 模块", "试配调整", "UHPC 试配框架页", "已存在页面入口和说明文案，但未接入 UHPC 工作性/强度/校正实际业务逻辑。", "全员", "/adapt/uhpc", "无", "部分完成", "审慎", "可按页面框架而非完整业务功能确认"],
]


PAGE_HEADERS = ["序号", "页面名称", "路由", "页面用途", "主要交互/能力", "访问角色", "当前状态", "备注"]

PAGE_ROWS = [
    [1, "登录页", "/login", "系统登录入口", "表单校验、登录提示、写入登录态", "公共", "已完成", "登录后跳转首页"],
    [2, "首页", "/", "系统首页与产品概览", "功能卡片展示、默认密码提醒、开始计算入口", "全员", "已完成", "业务首页"],
    [3, "项目管理", "/projects", "项目列表页", "搜索项目、新建项目、删除项目、进入详情", "全员/管理员", "已完成", "分页展示"],
    [4, "项目详情", "/projects/:id", "项目详情与记录归档页", "查看项目要求、记录列表、快捷新建配比/试配", "全员/管理员", "已完成", "项目维度工作台"],
    [5, "高性能混凝土配比计算", "/calc/hpc", "HPC 正式配比设计", "项目选择、五步计算、保存记录、右侧汇总、历史记录", "全员", "已完成", "核心交付页面"],
    [6, "高性能试配", "/adapt/hpc", "HPC 试配调整", "工作性实验、强度实验、配合比校正与确认", "全员", "已完成", "核心交付页面"],
    [7, "超高性能混凝土配比计算", "/calc/uhpc", "UHPC 页面入口", "当前仅展示开发中提示", "全员", "未完成/占位", "无实际计算能力"],
    [8, "超高性能试配", "/adapt/uhpc", "UHPC 试配页面框架", "入口、说明文案、返回 UHPC 配比页按钮", "全员", "部分完成", "暂无实际实验计算逻辑"],
    [9, "个人中心", "/profile", "个人资料与密码管理", "查看/修改邮箱手机号、修改密码", "全员", "已完成", "修改密码后重新登录"],
    [10, "用户管理", "/settings/users", "管理员账号管理", "搜索用户、新增用户、重置密码、删除用户", "管理员", "已完成", "仅管理员可见"],
    [11, "关于页", "/about", "系统说明页", "技术栈和计算流程展示", "全员", "已完成", "可作为客户说明材料"],
]


API_HEADERS = ["序号", "Method", "Path", "所属模块", "功能说明", "前端使用情况", "交付状态", "备注"]

API_ROWS = [
    [1, "GET", "/", "系统支撑", "根接口，返回系统运行状态与 docs 地址。", "调试/部署使用", "已完成", "非客户业务功能"],
    [2, "POST", "/api/auth/login", "认证", "用户登录，返回 token、用户名、管理员标记和 must_reset。", "当前前端使用", "已完成", "登录页调用"],
    [3, "POST", "/api/auth/logout", "认证", "用户退出登录，注销服务端 session。", "当前前端使用", "已完成", "全局退出登录调用"],
    [4, "POST", "/api/auth/reset-password", "认证", "首次登录后修改默认密码。", "后端已提供，当前前端未单独暴露页面", "已完成", "可作为后续增强入口"],
    [5, "GET", "/api/auth/profile", "个人中心", "获取当前登录用户资料和 must_reset 状态。", "当前前端使用", "已完成", "首页/个人中心使用"],
    [6, "PUT", "/api/auth/profile", "个人中心", "更新当前用户邮箱和手机号。", "当前前端使用", "已完成", "个人中心提交"],
    [7, "POST", "/api/auth/change-password", "个人中心", "修改当前用户密码。", "当前前端使用", "已完成", "个人中心提交"],
    [8, "GET", "/api/auth/users", "用户管理", "获取用户列表。", "当前前端使用", "已完成", "管理员能力"],
    [9, "POST", "/api/auth/users", "用户管理", "创建用户。", "当前前端使用", "已完成", "管理员能力"],
    [10, "POST", "/api/auth/users/{target_username}/reset-password", "用户管理", "管理员重置任意用户密码。", "当前前端使用", "已完成", "管理员能力"],
    [11, "DELETE", "/api/auth/users/{target_username}", "用户管理", "删除普通用户。", "当前前端使用", "已完成", "管理员能力"],
    [12, "GET", "/api/projects", "项目管理", "分页查询项目列表，支持搜索。", "当前前端使用", "已完成", "项目页和 HPC 页使用"],
    [13, "POST", "/api/projects", "项目管理", "创建项目。", "当前前端使用", "已完成", "项目列表页使用"],
    [14, "GET", "/api/projects/{project_id}", "项目管理", "查询项目详情。", "当前前端使用", "已完成", "项目详情页使用"],
    [15, "PUT", "/api/projects/{project_id}", "项目管理", "更新项目。", "当前前端使用", "已完成", "项目详情页使用"],
    [16, "DELETE", "/api/projects/{project_id}", "项目管理", "删除项目。", "当前前端使用", "已完成", "列表页/详情页使用"],
    [17, "GET", "/api/projects/{project_id}/records", "项目管理", "查询项目下关联配比记录。", "当前前端使用", "已完成", "项目详情页和 HPC 页使用"],
    [18, "POST", "/api/records", "记录管理", "保存或更新配比记录。", "当前前端使用", "已完成", "HPC 最终保存使用"],
    [19, "GET", "/api/records", "记录管理", "分页查询历史记录，支持项目和关键字过滤。", "当前前端使用", "已完成", "右侧历史记录使用"],
    [20, "DELETE", "/api/records/{record_id}", "记录管理", "删除指定记录。", "当前前端使用", "已完成", "项目详情/HPC 历史使用"],
    [21, "POST", "/api/water-binder", "HPC 配比计算", "计算配制强度和水胶比。", "当前前端使用", "已完成", "水胶比标签页使用"],
    [22, "POST", "/api/fit-coefficients", "HPC 配比计算", "根据 CSV 文本拟合回归系数。", "后端已提供，当前页面主要通过上传接口调用", "已完成", "可供后续文本导入使用"],
    [23, "POST", "/api/aggregate", "HPC 配比计算", "计算粗细骨料用量。", "当前前端使用", "已完成", "骨料标签页使用"],
    [24, "POST", "/api/binder", "HPC 配比计算", "计算胶材密度、胶凝总量和各组分用量。", "当前前端使用", "已完成", "胶凝材料标签页使用"],
    [25, "POST", "/api/water-admixture", "HPC 配比计算", "计算用水量与外加剂用量。", "当前前端使用", "已完成", "水和外加剂标签页使用"],
    [26, "POST", "/api/upload-fit", "HPC 配比计算", "上传 Excel/CSV 文件并拟合回归系数。", "当前前端使用", "已完成", "上传拟合按钮使用"],
    [27, "POST", "/api/adapt", "试配调整", "独立适配调整接口。", "保留/未接前端", "保留/未接前端", "当前 HPC 试配已改用 /api/hpc-trial"],
    [28, "POST", "/api/hpc-trial", "HPC 试配调整", "统一计算工作性实验、强度实验、配合比校正与确认结果。", "当前前端使用", "已完成", "高性能试配核心接口"],
]


BOUNDARY_HEADERS = ["序号", "模块/事项", "当前情况", "对客户影响", "建议对外表述", "本次结算建议"]

BOUNDARY_ROWS = [
    [1, "UHPC 配比计算", "页面已存在，但仅为开发中占位页，未接入公式、保存、历史和项目联动。", "客户无法完成 UHPC 正式配比计算。", "建议明确说明为“已预留入口，业务逻辑待后续开发”。", "否"],
    [2, "UHPC 试配调整", "页面入口和文案框架已存在，但工作性实验、强度实验、校正逻辑尚未实现。", "客户无法完成 UHPC 试配流程。", "建议说明为“已完成页面框架，未交付完整业务能力”。", "审慎"],
    [3, "独立 /api/adapt 接口", "后端保留该接口，但当前前端未调用，现行 HPC 试配通过 /api/hpc-trial 统一计算。", "不属于当前客户可直接使用的前台能力。", "可作为内部保留接口说明，不建议单独计费。", "不单独计费"],
    [4, "项目详情中的 UHPC 快捷入口", "项目详情页已预留 UHPC 新建入口，但落地页面仍是占位/框架状态。", "客户点击后只能进入占位页面。", "建议在验收时说明“入口已预留，功能待开发”。", "审慎"],
]


CLIENT_FEATURE_HEADERS = [
    "序号",
    "功能模块",
    "功能名称",
    "客户可见交付结果",
    "交付状态",
    "本次结算建议",
    "备注",
]

CLIENT_FEATURE_ROWS = [
    [1, "账号与权限", "登录/退出", "支持账号登录系统、保持登录状态并安全退出。", "已完成", "建议纳入", "基础使用入口"],
    [2, "账号与权限", "个人中心", "支持查看和维护个人资料，并修改登录密码。", "已完成", "建议纳入", "提升账号安全和日常维护效率"],
    [3, "账号与权限", "用户管理", "管理员可新增用户、重置密码、删除普通用户。", "已完成", "建议纳入", "管理员功能"],
    [4, "项目管理", "项目档案管理", "支持新建、搜索、查看、编辑、删除项目。", "已完成", "建议纳入", "项目编号唯一"],
    [5, "项目管理", "项目关联记录", "每个项目下可查看并管理对应的配比记录。", "已完成", "建议纳入", "支持按项目归档"],
    [6, "HPC 正式配比", "水胶比计算", "支持输入强度参数自动计算配制强度和水胶比。", "已完成", "建议纳入", "支持默认系数和自定义系数"],
    [7, "HPC 正式配比", "回归系数拟合", "支持导入 Excel/CSV 数据拟合回归系数。", "已完成", "建议纳入", "适合结合试验数据修正参数"],
    [8, "HPC 正式配比", "砂率选取", "支持通过参考表选择并确认砂率。", "已完成", "建议纳入", "与后续骨料计算联动"],
    [9, "HPC 正式配比", "骨料用量计算", "支持计算粗骨料和细骨料用量。", "已完成", "建议纳入", "自动引用砂率结果"],
    [10, "HPC 正式配比", "胶凝材料计算", "支持计算胶材总量，并拆分水泥、粉煤灰、矿粉、微珠、硅灰用量。", "已完成", "建议纳入", "覆盖多种掺合料场景"],
    [11, "HPC 正式配比", "用水与外加剂计算", "支持计算用水量、外加剂用量和每方材料合计。", "已完成", "建议纳入", "可直接保存为正式配比记录"],
    [12, "HPC 正式配比", "历史记录与汇总", "支持查看历史记录、载入记录、删除记录和右侧结果汇总。", "已完成", "建议纳入", "便于复用历史配比"],
    [13, "HPC 试配调整", "工作性实验", "支持在保持 W/B 不变的前提下调整胶材、砂率、外加剂，并生成试拌用量。", "已完成", "建议纳入", "支持记录实测工作性描述"],
    [14, "HPC 试配调整", "强度实验", "支持生成三组试验配比、录入强度结果并自动回归推荐水胶比。", "已完成", "建议纳入", "包含回归结果和图示"],
    [15, "HPC 试配调整", "配合比校正与确认", "支持依据推荐水胶比和表观密度校正实验室配合比。", "已完成", "建议纳入", "形成试配闭环"],
    [16, "系统页面", "首页与说明页", "支持首页展示、快速开始计算和计算流程说明。", "已完成", "建议纳入", "适合作为系统介绍页"],
    [17, "UHPC 模块", "UHPC 配比计算", "当前仅提供页面入口和开发中提示，尚未形成完整计算功能。", "未完成/占位", "暂不建议", "建议作为后续开发范围"],
    [18, "UHPC 模块", "UHPC 试配调整", "当前已完成页面框架和入口，但未接入实际试配逻辑。", "部分完成", "审慎", "建议明确为后续开发项"],
]


CLIENT_SCOPE_HEADERS = [
    "序号",
    "结算模块",
    "客户可见交付结果",
    "本次结算建议",
    "备注",
]

CLIENT_SCOPE_ROWS = [
    [1, "账号与权限", "支持登录、退出、个人资料维护和管理员账号管理。", "建议纳入", "已可直接投入使用"],
    [2, "项目管理", "支持项目创建、查询、编辑、删除，以及项目维度的配比记录归档。", "建议纳入", "已形成项目管理闭环"],
    [3, "HPC 正式配比设计", "支持五步正式配比计算、参数拟合、结果汇总与保存。", "建议纳入", "属于核心业务功能"],
    [4, "HPC 试配调整", "支持工作性实验、强度实验回归推荐和配合比校正。", "建议纳入", "已形成试配闭环"],
    [5, "记录管理", "支持历史记录查询、载入和删除。", "建议纳入", "便于数据追溯和复用"],
    [6, "首页与说明页", "支持产品说明、快速入口和流程展示。", "建议纳入", "可作为客户培训辅助页面"],
    [7, "UHPC 配比计算", "当前仅为开发中页面，未交付完整业务逻辑。", "暂不建议", "不建议按完整功能结算"],
    [8, "UHPC 试配调整", "当前仅完成页面框架和入口接入。", "审慎", "若要计费，建议只按页面框架确认"],
]


def set_base_style(cell, *, bold: bool = False, font_color: str = "1F1F1F", fill: PatternFill | None = None) -> None:
    cell.font = Font(name=FONT_NAME, size=10, bold=bold, color=font_color)
    cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    cell.border = TABLE_BORDER
    if fill is not None:
        cell.fill = fill


def merge_and_style(ws, start_row: int, start_col: int, end_row: int, end_col: int, value: str, *, fill: PatternFill, font_size: int, bold: bool = True, font_color: str = "1F1F1F") -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(start_row, start_col, value)
    cell.font = Font(name=FONT_NAME, size=font_size, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = fill
    cell.border = TABLE_BORDER
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for item in row:
            item.border = TABLE_BORDER
            if item.coordinate != cell.coordinate:
                item.fill = fill


def set_column_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def style_data_table(ws, header_row: int, data_start_row: int, data_end_row: int, widths: list[int], *, status_col: int | None = None, settlement_col: int | None = None) -> None:
    set_column_widths(ws, widths)
    ws.freeze_panes = f"A{data_start_row}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(widths))}{data_end_row}"

    for cell in ws[header_row]:
        set_base_style(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index in range(data_start_row, data_end_row + 1):
        fill = ALT_FILL if (row_index - data_start_row) % 2 else None
        for col_index in range(1, len(widths) + 1):
            cell = ws.cell(row_index, col_index)
            set_base_style(cell, fill=fill)
            if col_index == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        if status_col is not None:
            status_cell = ws.cell(row_index, status_col)
            status_fill = STATUS_FILLS.get(str(status_cell.value or ""))
            if status_fill is not None:
                set_base_style(status_cell, fill=status_fill)
                status_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if settlement_col is not None:
            settlement_cell = ws.cell(row_index, settlement_col)
            settlement_fill = SETTLEMENT_FILLS.get(str(settlement_cell.value or ""))
            if settlement_fill is not None:
                set_base_style(settlement_cell, fill=settlement_fill)
                settlement_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_table(ws, start_row: int, headers: list[str], rows: list[list[object]], widths: list[int], *, status_col: int | None = None, settlement_col: int | None = None) -> None:
    header_row = start_row
    data_start_row = start_row + 1

    for column, header in enumerate(headers, start=1):
        ws.cell(header_row, column, header)

    for row_offset, values in enumerate(rows, start=0):
        for column, value in enumerate(values, start=1):
            ws.cell(data_start_row + row_offset, column, value)

    style_data_table(
        ws,
        header_row,
        data_start_row,
        data_start_row + len(rows) - 1,
        widths,
        status_col=status_col,
        settlement_col=settlement_col,
    )


def build_overview_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("概览")
    ws.sheet_view.showGridLines = False

    merge_and_style(
        ws,
        1,
        1,
        1,
        6,
        "水泥配比计算器 客户结算功能需求文档",
        fill=TITLE_FILL,
        font_size=16,
        font_color="FFFFFF",
    )
    merge_and_style(
        ws,
        2,
        1,
        2,
        6,
        "根据当前代码实际实现整理，适合向客户确认交付范围与本次结算边界。",
        fill=SUBTITLE_FILL,
        font_size=10,
        bold=False,
    )

    merge_and_style(ws, 4, 1, 4, 6, "一、文档说明", fill=SECTION_FILL, font_size=11)

    overview_rows = [
        ["系统名称", "水泥配比计算器（Vue 3 + FastAPI）"],
        ["整理日期", date.today().isoformat()],
        ["文档用途", "用于向客户说明当前版本已落地功能，并作为结算前的功能确认清单。"],
        ["当前建议结算范围", "登录与权限、项目管理、HPC 配比计算全流程、HPC 试配调整全流程、记录管理、用户管理、个人中心、首页与说明页。"],
        ["暂不建议按完整功能结算范围", "UHPC 配比计算、UHPC 试配调整；当前仅完成入口、占位页或页面框架。"],
        ["对客户的建议表述", "建议以“功能清单”中交付状态=已完成且建议纳入本次结算=是 的条目作为本次结算主体。"],
    ]

    ws.cell(5, 1, "字段")
    ws.cell(5, 2, "内容")
    for cell in ws[5][0:2]:
        set_base_style(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = 6
    for key, value in overview_rows:
        ws.cell(current_row, 1, key)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
        ws.cell(current_row, 2, value)
        set_base_style(ws.cell(current_row, 1), bold=True)
        set_base_style(ws.cell(current_row, 2))
        for merged_col in range(3, 7):
            set_base_style(ws.cell(current_row, merged_col))
        current_row += 1

    merge_and_style(ws, 13, 1, 13, 6, "二、交付统计", fill=SECTION_FILL, font_size=11)

    completed_features = sum(1 for row in FEATURE_ROWS if row[8] == "已完成")
    partial_features = sum(1 for row in FEATURE_ROWS if row[8] == "部分完成")
    unfinished_features = sum(1 for row in FEATURE_ROWS if row[8] == "未完成/占位")
    completed_pages = sum(1 for row in PAGE_ROWS if row[6] == "已完成")
    partial_pages = sum(1 for row in PAGE_ROWS if row[6] == "部分完成")
    unfinished_pages = sum(1 for row in PAGE_ROWS if row[6] == "未完成/占位")
    active_apis = sum(1 for row in API_ROWS if row[6] == "已完成")
    reserved_apis = sum(1 for row in API_ROWS if row[6] == "保留/未接前端")

    stats_headers = ["指标", "数量", "说明"]
    stats_rows = [
        ["已梳理功能项", len(FEATURE_ROWS), "包含已完成、部分完成和占位项"],
        ["已完成功能项", completed_features, "建议作为本次结算主体"],
        ["部分完成功能项", partial_features, "需与客户说明仅交付框架或部分能力"],
        ["未完成/占位功能项", unfinished_features, "不建议按完整功能结算"],
        ["页面数量", len(PAGE_ROWS), f"其中已完成 {completed_pages} 个，部分完成 {partial_pages} 个，未完成 {unfinished_pages} 个"],
        ["接口数量", len(API_ROWS), f"其中当前前端实际使用/可用 {active_apis} 个，保留接口 {reserved_apis} 个"],
    ]
    write_table(ws, 14, stats_headers, stats_rows, [22, 12, 68])

    merge_and_style(ws, 22, 1, 22, 6, "三、发客户时可直接说明的结论", fill=SECTION_FILL, font_size=11)

    conclusion_rows = [
        ["1", "本系统当前已形成完整的 HPC 配比设计与 HPC 试配调整闭环，可支撑项目创建、配比计算、试配验证、记录归档和权限管理。"],
        ["2", "UHPC 两个页面目前已完成导航接入与页面占位/框架，但尚未形成可直接使用的完整业务能力。"],
        ["3", "若本次以“当前可用功能”为结算依据，建议优先以功能清单中状态为“已完成”的条目进行确认。"],
        ["4", "如果合同中包含 UHPC 完整计算与试配能力，需要明确说明该部分仍属于后续开发范围。"],
    ]

    ws.cell(23, 1, "序号")
    ws.cell(23, 2, "建议说明")
    for cell in ws[23][0:2]:
        set_base_style(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, row in enumerate(conclusion_rows, start=24):
        ws.cell(row_index, 1, row[0])
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=6)
        ws.cell(row_index, 2, row[1])
        set_base_style(ws.cell(row_index, 1), bold=True)
        set_base_style(ws.cell(row_index, 2))
        for merged_col in range(3, 7):
            set_base_style(ws.cell(row_index, merged_col))

    set_column_widths(ws, [18, 24, 14, 14, 14, 14])
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24


def build_feature_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("功能清单")
    ws.sheet_view.showGridLines = False
    merge_and_style(ws, 1, 1, 1, len(FEATURE_HEADERS), "功能清单", fill=TITLE_FILL, font_size=14, font_color="FFFFFF")
    merge_and_style(ws, 2, 1, 2, len(FEATURE_HEADERS), "本表适合作为客户确认和结算对账的主体清单。", fill=SUBTITLE_FILL, font_size=10, bold=False)
    write_table(ws, 4, FEATURE_HEADERS, FEATURE_ROWS, [8, 14, 14, 18, 40, 12, 18, 26, 12, 14, 24], status_col=9, settlement_col=10)


def build_page_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("页面清单")
    ws.sheet_view.showGridLines = False
    merge_and_style(ws, 1, 1, 1, len(PAGE_HEADERS), "页面清单", fill=TITLE_FILL, font_size=14, font_color="FFFFFF")
    merge_and_style(ws, 2, 1, 2, len(PAGE_HEADERS), "从客户可见页面角度说明当前系统范围。", fill=SUBTITLE_FILL, font_size=10, bold=False)
    write_table(ws, 4, PAGE_HEADERS, PAGE_ROWS, [8, 20, 20, 20, 34, 12, 12, 24], status_col=7)


def build_api_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("接口清单")
    ws.sheet_view.showGridLines = False
    merge_and_style(ws, 1, 1, 1, len(API_HEADERS), "接口清单", fill=TITLE_FILL, font_size=14, font_color="FFFFFF")
    merge_and_style(ws, 2, 1, 2, len(API_HEADERS), "用于说明后台已实现能力，以及哪些接口是当前前端真实在用。", fill=SUBTITLE_FILL, font_size=10, bold=False)
    write_table(ws, 4, API_HEADERS, API_ROWS, [8, 10, 34, 16, 34, 20, 14, 24], status_col=7)


def build_boundary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("待开发与边界")
    ws.sheet_view.showGridLines = False
    merge_and_style(ws, 1, 1, 1, len(BOUNDARY_HEADERS), "待开发与边界", fill=TITLE_FILL, font_size=14, font_color="FFFFFF")
    merge_and_style(ws, 2, 1, 2, len(BOUNDARY_HEADERS), "本表用于避免客户把占位页或保留接口误解为已交付的完整功能。", fill=SUBTITLE_FILL, font_size=10, bold=False)
    write_table(ws, 4, BOUNDARY_HEADERS, BOUNDARY_ROWS, [8, 18, 34, 24, 34, 16], settlement_col=6)


def build_client_overview_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("客户说明")
    ws.sheet_view.showGridLines = False

    merge_and_style(
        ws,
        1,
        1,
        1,
        5,
        "水泥配比计算器 客户简版功能确认单",
        fill=TITLE_FILL,
        font_size=16,
        font_color="FFFFFF",
    )
    merge_and_style(
        ws,
        2,
        1,
        2,
        5,
        "用于客户快速确认当前阶段已交付的主要功能与本次建议结算范围。",
        fill=SUBTITLE_FILL,
        font_size=10,
        bold=False,
    )

    merge_and_style(ws, 4, 1, 4, 5, "一、文档用途", fill=SECTION_FILL, font_size=11)
    usage_rows = [
        ["文档用途", "帮助客户从业务视角快速确认当前系统已交付范围，不展开接口和底层技术细节。"],
        ["适用场景", "用于阶段性验收、结算确认、客户内部汇报。"],
        ["建议阅读方式", "优先看“客户功能清单”和“结算范围建议”两个工作表。"],
    ]

    ws.cell(5, 1, "字段")
    ws.cell(5, 2, "内容")
    for cell in ws[5][0:2]:
        set_base_style(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = 6
    for key, value in usage_rows:
        ws.cell(current_row, 1, key)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        ws.cell(current_row, 2, value)
        set_base_style(ws.cell(current_row, 1), bold=True)
        set_base_style(ws.cell(current_row, 2))
        for merged_col in range(3, 6):
            set_base_style(ws.cell(current_row, merged_col))
        current_row += 1

    merge_and_style(ws, 10, 1, 10, 5, "二、本阶段主要已交付结果", fill=SECTION_FILL, font_size=11)
    delivery_headers = ["序号", "已交付结果"]
    delivery_rows = [
        [1, "已完成账号登录、个人中心、管理员用户管理能力。"],
        [2, "已完成项目管理，并支持按项目归档配比记录。"],
        [3, "已完成高性能混凝土（HPC）正式配比计算全流程。"],
        [4, "已完成高性能混凝土（HPC）试配调整全流程。"],
        [5, "已完成历史记录查询、载入、删除和结果汇总展示。"],
    ]
    write_table(ws, 11, delivery_headers, delivery_rows, [10, 78])

    merge_and_style(ws, 18, 1, 18, 5, "三、需单独说明的边界", fill=SECTION_FILL, font_size=11)
    boundary_headers = ["序号", "说明事项"]
    boundary_rows = [
        [1, "UHPC 配比计算当前仅为开发中页面，尚未形成完整业务功能。"],
        [2, "UHPC 试配调整当前仅完成页面框架和入口接入，尚未形成完整试配逻辑。"],
        [3, "因此本次建议以 HPC 相关完整功能为主要结算范围。"],
    ]
    write_table(ws, 19, boundary_headers, boundary_rows, [10, 78])


def build_client_feature_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("客户功能清单")
    ws.sheet_view.showGridLines = False
    merge_and_style(ws, 1, 1, 1, len(CLIENT_FEATURE_HEADERS), "客户功能清单", fill=TITLE_FILL, font_size=14, font_color="FFFFFF")
    merge_and_style(ws, 2, 1, 2, len(CLIENT_FEATURE_HEADERS), "仅保留客户易理解的功能项，不展示接口、路由和过多技术细节。", fill=SUBTITLE_FILL, font_size=10, bold=False)
    write_table(ws, 4, CLIENT_FEATURE_HEADERS, CLIENT_FEATURE_ROWS, [8, 16, 18, 42, 12, 14, 24], status_col=5, settlement_col=6)


def build_client_scope_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("结算范围建议")
    ws.sheet_view.showGridLines = False
    merge_and_style(ws, 1, 1, 1, len(CLIENT_SCOPE_HEADERS), "结算范围建议", fill=TITLE_FILL, font_size=14, font_color="FFFFFF")
    merge_and_style(ws, 2, 1, 2, len(CLIENT_SCOPE_HEADERS), "适合直接给客户确认哪些模块建议纳入本次结算。", fill=SUBTITLE_FILL, font_size=10, bold=False)
    write_table(ws, 4, CLIENT_SCOPE_HEADERS, CLIENT_SCOPE_ROWS, [8, 18, 42, 14, 24], settlement_col=4)


def create_full_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "水泥配比计算器客户结算功能需求文档"
    wb.properties.subject = "客户结算功能清单"
    wb.properties.creator = "GitHub Copilot"
    wb.properties.description = "根据当前项目代码实际实现生成的客户结算功能清单。"

    build_overview_sheet(wb)
    build_feature_sheet(wb)
    build_page_sheet(wb)
    build_api_sheet(wb)
    build_boundary_sheet(wb)
    return wb


def create_client_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "水泥配比计算器客户简版功能确认单"
    wb.properties.subject = "客户简版功能清单"
    wb.properties.creator = "GitHub Copilot"
    wb.properties.description = "根据当前项目代码实际实现生成的客户简版功能确认单。"

    build_client_overview_sheet(wb)
    build_client_feature_sheet(wb)
    build_client_scope_sheet(wb)
    return wb


def main() -> None:
    output_dir = Path(__file__).resolve().parent
    full_output_path = output_dir / f"shuini_calculator_requirement_doc_{date.today().isoformat()}.xlsx"
    client_output_path = output_dir / f"shuini_calculator_client_brief_{date.today().isoformat()}.xlsx"

    full_wb = create_full_workbook()
    full_wb.save(full_output_path)

    client_wb = create_client_workbook()
    client_wb.save(client_output_path)

    print(full_output_path)
    print(client_output_path)


if __name__ == "__main__":
    main()