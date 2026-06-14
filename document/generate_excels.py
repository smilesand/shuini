"""
生成功能清单 Excel 和 验收表格 Excel
运行: python document/generate_excels.py
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from copy import copy

OUT_DIR = "document"

# ── 样式常量 ──
DARK_BLUE = "1E3C72"
MID_BLUE = "2A5298"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F7FA"
GREEN = "2E7D32"
LIGHT_GREEN = "E8F5E9"
RED = "C62828"

FONT_TITLE = Font(name="微软雅黑", size=16, bold=True, color=DARK_BLUE)
FONT_HEADER = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
FONT_SECTION = Font(name="微软雅黑", size=11, bold=True, color=DARK_BLUE)
FONT_NORMAL = Font(name="微软雅黑", size=10, color="333333")
FONT_SMALL = Font(name="微软雅黑", size=9, color="666666")

FILL_HEADER = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
FILL_SECTION = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
FILL_ROW_EVEN = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_GREEN = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_range(ws, row_start, row_end, col_start, col_end, font=None, fill=None, alignment=None, border=None):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border


# ═══════════════════════════════════════════
# 📊 功能清单 Excel
# ═══════════════════════════════════════════

FEATURES = [
    # (模块, 一级菜单, 功能名称, 功能详细说明, 工时/h)
    ("系统基础", "—", "用户认证与JWT鉴权",
     "基于JWT Token的登录认证机制，支持Token过期自动跳转登录页。前端路由守卫拦截未登录用户，后端API层统一校验Token有效性。",
     2.5),
    ("系统基础", "—", "默认管理员初始化",
     "首次启动时自动创建默认管理员账号(admin/123456)，并标记密码需重置。管理员账号不可通过常规删除接口删除。",
     0.5),
    ("系统基础", "—", "密码安全策略",
     "首次登录检测默认密码并主动提示修改；个人中心支持自主修改密码（需原密码验证）；管理员可强制重置任意用户密码；密码使用SHA-256加盐哈希存储。",
     1),
    ("系统基础", "—", "用户角色与权限控制",
     "区分管理员(admin)与普通用户(user)两种角色。管理员可见「系统设置」菜单进行用户管理；普通用户仅可见业务功能菜单。后端API层对管理员接口做二次鉴权。",
     1),

    ("首页总览", "首页", "统计卡片展示",
     "首页展示4张统计卡片：项目总数、HPC计算记录数、UHPC计算记录数、试配记录总数。数据通过后端统计接口实时获取。",
     1),
    ("首页总览", "首页", "快捷操作面板",
     "一行五格快捷入口：HPC配比计算、UHPC配比计算、HPC试配调整、UHPC试配调整、项目管理。点击直接跳转对应页面。",
     1),
    ("首页总览", "首页", "最近项目与最近记录",
     "最近项目展示5条（含编号、名称、记录数，可点击进入详情）；最近配合记录展示10条（含类别标签、名称、时间，可点击载入）。",
     1),
    ("首页总览", "首页", "顶部Banner与问候语",
     "自适应时段问候语(早上好/中午好/下午好/晚上好)+用户名+日期星期；品牌Banner展示CRRC标识；「开始计算」「项目管理」快捷按钮。",
     0.5),

    ("项目管理", "项目管理", "项目列表与分页搜索",
     "表格展示全部项目(编号/名称/要求/创建人/配比数/时间)，支持分页浏览(每页10条)；按项目编号或名称关键字搜索过滤。",
     1),
    ("项目管理", "项目管理", "新建项目",
     "弹窗表单创建项目：项目编号(必填)、项目名称(必填)、项目要求(可选，支持多行文本)。",
     0.5),
    ("项目管理", "项目管理", "编辑与删除项目",
     "弹窗编辑项目信息；软删除项目（项目及关联记录一并移入回收站），删除前弹窗二次确认。",
     1),
    ("项目管理", "项目详情", "项目信息与记录管理",
     "展示项目完整信息；表格展示该项目下所有配比记录(含全部材料用量列与关键参数列)；支持新建HPC/UHPC配比并自动关联项目；支持载入和删除记录。",
     1.5),
    ("项目管理", "项目详情", "记录数据表格(多列)",
     "12列材料与参数表格：水泥/粉煤灰/矿粉/微珠/硅灰/细骨料/粗骨料/水/外加剂/钢纤维，胶材总量/水胶比/砂率/总量。支持前后端分页与排序。",
     1),

    ("HPC配比计算", "配比计算→HPC", "项目选择与历史记录",
     "进入HPC计算页后首先展示项目选择下拉框与该项目下已有HPC记录列表，支持新建配比或载入历史记录。支持切换项目。",
     1),
    ("HPC配比计算", "配比计算→HPC", "水胶比页签（鲍罗米公式）",
     "输入强度等级fcu,k和胶材28d强度fb，系统基于鲍罗米公式自动计算配制强度fcu,0和水胶比W/B。支持手动输入或自动估算fb(输入fce,g与γc)；支持上传CSV/Excel试验数据回归拟合系数αa/αb。",
     3.5),
    ("HPC配比计算", "配比计算→HPC", "砂率选取页签（参考表交互）",
     "提供标准砂率参考表(强度等级×骨料粒径)，单击行列交叉高亮推荐范围。用户手动输入目标砂率后系统确认并传递到下游计算。",
     2),
    ("HPC配比计算", "配比计算→HPC", "骨料用量页签（体积法）",
     "体积法计算：输入粗骨料绝对体积Vg、粗骨料密度ρg、细骨料密度ρs，系统自动计算粗骨料mg和细骨料ms用量。附带Vg推荐参考表。",
     2),
    ("HPC配比计算", "配比计算→HPC", "胶凝材料页签（浆体体积法）",
     "浆体体积法计算：支持四种掺合料(粉煤灰/矿粉/微珠/硅灰)的质量分数与密度输入，计算各组分每方用量mc/m1/m2/m3/m4及胶材总量mb。含气量参数可调。",
     3),
    ("HPC配比计算", "配比计算→HPC", "水和外加剂页签与汇总",
     "根据mb×W/B计算用水量mw，输入外加剂掺量百分比计算外加剂用量ma，汇总全部配比结果并展示最终配比表（全部材料每方用量及总量）。",
     2),
    ("HPC配比计算", "配比计算→HPC", "分步式页签流程与侧栏",
     "5个页签顺序引导（水胶比→砂率→骨料→胶材→水和外加剂），每步依赖前序结果。左侧侧栏实时展示关键参数(W/B、βs)及状态指示灯。支持历史记录搜索/分页/载入/删除。",
     1.5),
    ("HPC配比计算", "配比计算→HPC", "保存配比记录",
     "输入配比名称后保存完整计算数据到数据库。记录关联项目，包含全部输入参数和计算结果。支持后续载入复用。",
     0.5),

    ("UHPC配比计算", "配比计算→UHPC", "项目选择与历史记录",
     "与HPC计算页同构的项目选择+历史记录界面，专用于UHPC配比。支持新建配比或载入历史记录。",
     1),
    ("UHPC配比计算", "配比计算→UHPC", "水胶比页签（推荐表）",
     "提供UHPC推荐强度等级与W/B对照表，点击推荐行自动带入推荐水胶比；支持手动输入外加剂掺量(空值时自动补默认值)。",
     2),
    ("UHPC配比计算", "配比计算→UHPC", "砂胶比与钢纤维页签",
     "砂胶比S/B输入后计算砂体积占比和关键体积参数；钢纤维体积掺量与密度输入后自动计算钢纤维质量并调整配比。",
     2.5),
    ("UHPC配比计算", "配比计算→UHPC", "胶凝材料比例页签（安德森模型）",
     "基于修正安德森模型的粒径堆积计算：输入各材料粒径参数(D10/D50/D90)、微粉参数与密度，自动计算胶凝材料体积比例→初始质量比例→修正质量比例→每方用量。不同强度等级显示对应的微粉系数参考提示。",
     4),
    ("UHPC配比计算", "配比计算→UHPC", "分步式页签流程与侧栏",
     "4个页签顺序引导（水胶比→砂胶比→钢纤维→胶凝材料比例）。左侧侧栏实时汇总UHPC关键参数；保存配比记录。",
     1.5),

    ("HPC试配调整", "试配调整→HPC", "工作性实验",
     "在基准配比基础上调整胶材用量增减量Δmb、砂率增减量Δβs、外加剂掺量增减量Δα。系统自动(500ms防抖)重算试拌配合比和试拌用量。附带HPC工作性评价参考指标。",
     3),
    ("HPC试配调整", "试配调整→HPC", "强度实验与回归分析",
     "生成+Δ/-Δ水胶比试验组配比，录入各组实测强度后系统执行最小二乘回归分析，输出回归系数a/b、相关系数R²、推荐水胶比和预测强度，并绘制回归曲线图表。",
     3.5),
    ("HPC试配调整", "试配调整→HPC", "配合比校正与记录保存",
     "输入拌合物表观密度实测值，自动计算校正系数δ并生成校正后的实验室配合比(每方各材料用量)。试配快照以独立数据结构附着在主记录中保存。页面顶部展示关联的项目名和记录名。",
     2),

    ("UHPC试配调整", "试配调整→UHPC", "工作性实验",
     "调整砂胶比S/B和外加剂掺量α，系统自动重新生成满足新参数并考虑体积约束的试拌配合比。",
     2),
    ("UHPC试配调整", "试配调整→UHPC", "强度实验与回归分析",
     "通过等质量替代策略生成+Δ/-Δ试验组，支持不同水胶比或硅灰掺量的调整方案。录入实测强度后进行二次回归分析生成推荐配比方案。",
     3.5),
    ("UHPC试配调整", "试配调整→UHPC", "配合比校正与确定",
     "提供三种校正基准选择（试拌/推荐水胶比/推荐硅灰用量）。输入实测表观密度后自动计算校正系数，生成最终UHPC实验室配合比。试配快照附着在主记录中保存。",
     2.5),

    ("计算书", "计算书", "项目选择与计算书生成",
     "左侧项目列表(支持搜索过滤)；点击项目展示记录列表；点击「生成计算书」弹出格式化HTML计算书（项目信息+全部计算参数键值表），支持浏览器打印/导出PDF。",
     2.5),

    ("全部配合记录", "全部记录", "记录列表与多条件筛选",
     "统一展示全部HPC/UHPC配比记录表格(12列)。支持按配比名称/项目名称搜索，按类别(HPC/UHPC)下拉筛选，实时前端分页。支持手动刷新数据。",
     1.5),
    ("全部配合记录", "全部记录", "载入历史记录",
     "点击「载入」按钮，自动跳转对应计算页面(HPC→/calc/hpc, UHPC→/calc/uhpc)并填入全部参数。",
     1),

    ("回收站", "回收站", "回收站管理（恢复/彻底删除）",
     "展示所有软删除项目和记录；按类型(全部/项目/记录)筛选/搜索；管理员视图查看全部用户数据；恢复项目时同步恢复关联记录，恢复记录时若项目在回收站则一并恢复；物理删除前弹窗二次警告。",
     2.5),

    ("个人中心", "个人中心", "个人资料与密码管理",
     "展示用户名(只读)、邮箱、手机号并支持修改保存。修改密码：原密码验证→新密码(4-100位)→确认密码一致性校验。修改成功后自动退出要求重新登录。",
     1.5),

    ("用户管理", "系统设置→用户管理", "用户管理（CRUD）",
     "表格展示所有用户（含角色/密码状态标签），支持搜索过滤。新增用户(默认密码123456)；管理员重置任意用户密码；删除普通用户(管理员不可删)。仅管理员可见该菜单，后端二次鉴权。",
     2.5),

    ("系统部署", "—", "前端生产构建与部署",
     "Vite生产模式打包：Tree-shaking、代码分割、资源压缩。输出dist目录可直接部署到Nginx。配置Nginx反向代理与SPA路由重写。",
     2),
    ("系统部署", "—", "后端PyInstaller打包",
     "使用PyInstaller将FastAPI后端打包为独立可执行文件(main)，内嵌SQLite驱动。支持Windows/Linux交叉构建shell/PowerShell脚本。",
     1.5),
    ("系统部署", "—", "数据库自动初始化与迁移",
     "首次启动时自动创建SQLite数据库及全部表结构，插入默认管理员账号。后续启动做schema兼容迁移。统一日志系统(文件+控制台，按日期轮转)。",
     1.5),

    ("API接口", "—", "RESTful API体系",
     "完整CRUD接口：认证(login/profile/password)、项目管理、配比记录管理、回收站管理、用户管理。Pydantic请求校验、统一异常处理中间件、SQLite外键约束。",
     3),
    ("API接口", "—", "科学计算引擎",
     "HPC计算引擎(鲍罗米公式/体积法/浆体体积法)；UHPC计算引擎(安德森模型/粒径堆积)；回归拟合引擎(最小二乘法)；试配计算引擎(水胶比调整/强度回归/密度校正)。NumPy/SciPy集成。",
     3.5),

    ("Bug修复与优化", "—", "Bug修复与边界测试",
     "修复开发与测试过程中发现的全部功能缺陷；覆盖空值输入、极端参数、并发操作等边界场景测试；前后端异常处理完善。",
     3),
    ("Bug修复与优化", "—", "性能优化",
     "前端：组件懒加载、防抖节流、虚拟列表、首屏加载优化。后端：SQLite查询优化、接口响应缓存、科学计算异步化。打包体积优化。",
     2),
    ("Bug修复与优化", "—", "UI/UX 打磨优化",
     "统一样式规范、响应式适配(1366×768及以上)、加载状态与空状态设计、操作反馈(成功/失败/确认弹窗)、品牌配色一致性、交互细节打磨。",
     2),
]


def generate_feature_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "功能清单"

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12

    # ── 标题行 ──
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "水泥配比计算系统 —— 功能清单"
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = "版本: 2026-06-02 | 文档类型: 功能交付清单"
    c.font = FONT_SMALL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # ── 表头 ──
    headers = ["序号", "功能模块", "菜单路径", "功能名称", "功能详细说明", "工时(h)", "备注"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 28

    # ── 数据行 ──
    row = 5
    last_module = ""
    seq = 0
    total_hours = 0

    for module, menu, name, desc, hours in FEATURES:
        seq += 1
        total_hours += hours

        is_section_first = (module != last_module)
        last_module = module

        values = [seq, module, menu, name, desc, hours, ""]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT if col_idx <= 5 else ALIGN_CENTER
            cell.border = THIN_BORDER

            if is_section_first and col_idx in (1, 2):
                cell.font = FONT_SECTION
                cell.fill = FILL_SECTION

        # 奇偶行底色
        if seq % 2 == 0:
            for col_idx in range(1, 8):
                if not ws.cell(row=row, column=col_idx).fill or ws.cell(row=row, column=col_idx).fill.start_color.index == "00000000":
                    ws.cell(row=row, column=col_idx).fill = FILL_ROW_EVEN

        ws.row_dimensions[row].height = 36
        row += 1

    # ── 合计行 ──
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value=f"合计: {len(FEATURES)} 项功能")
    c.font = Font(name="微软雅黑", size=11, bold=True, color=DARK_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center")
    for col_idx in range(1, 8):
        ws.cell(row=row, column=col_idx).border = THIN_BORDER

    c = ws.cell(row=row, column=6, value=total_hours)
    c.font = Font(name="微软雅黑", size=11, bold=True, color=DARK_BLUE)
    c.alignment = ALIGN_CENTER
    c.number_format = "0.0"

    c = ws.cell(row=row, column=7, value=f"≈ {total_hours/8:.1f} 人天")
    c.font = Font(name="微软雅黑", size=10, bold=True, color="666666")
    c.alignment = ALIGN_CENTER

    ws.row_dimensions[row].height = 30

    # ── 冻结窗格 ──
    ws.freeze_panes = "A5"

    # ── 打印设置 ──
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    path = f"{OUT_DIR}/功能清单_2026-06-02.xlsx"
    wb.save(path)
    print(f"✅ 功能清单已生成: {path}")
    print(f"   共 {len(FEATURES)} 项功能, 合计 {total_hours} 工时 (≈{total_hours/8:.1f}人天)")


# ═══════════════════════════════════════════
# 📋 验收表格 Excel
# ═══════════════════════════════════════════

ACCEPTANCE_ITEMS = [
    # (模块, 验收项, 验收标准, 权重)
    ("系统基础", "用户登录与JWT认证",
     "输入正确的用户名和密码后可成功登录；使用错误凭据时显示明确错误提示。Token过期后自动跳转登录页。",
     "高"),
    ("系统基础", "默认管理员初始化",
     "首次启动后自动创建admin账号(密码123456)；首次登录后提示修改密码。",
     "高"),
    ("系统基础", "密码安全",
     "个人中心可修改密码(需原密码验证)；管理员可在用户管理中重置任意用户密码；密码修改成功后自动退出要求重新登录。",
     "高"),
    ("系统基础", "角色与权限控制",
     "管理员可见「系统设置」菜单并进行用户管理；普通用户登录后不显示系统设置菜单；直接访问管理页面URL被拦截。",
     "高"),

    ("首页总览", "统计卡片数据准确",
     "首页4张统计卡片数值与数据库实际数据一致：项目总数、HPC/UHPC记录数、试配记录总数。",
     "中"),
    ("首页总览", "快捷操作跳转正确",
     "5个快捷卡片点击后均跳转到正确目标页面(HPC计算/UHPC计算/HPC试配/UHPC试配/项目管理)。",
     "中"),
    ("首页总览", "最近项目与最近记录",
     "最近项目显示5条，点击可进入详情；最近记录显示10条，点击可载入对应计算页。数据按时间倒序排列。",
     "低"),

    ("项目管理", "项目CRUD",
     "可正常创建项目(编号+名称+要求)；编辑项目信息并保存；搜索项目按关键词过滤；删除项目后进入回收站。",
     "高"),
    ("项目管理", "项目详情页记录管理",
     "项目详情页正确展示该项目下所有配比记录；可新建HPC/UHPC配比并自动关联项目；可载入和删除记录。",
     "高"),

    ("HPC配比计算", "水胶比计算",
     "输入fcu,k和fb后自动计算fcu,0和W/B；fb自动估算面板正常工作(输入fce,g和γc)；回归拟合文件上传后系数正确更新。",
     "高"),
    ("HPC配比计算", "砂率选取",
     "参考表行列点击后正确交叉高亮推荐范围；手动输入砂率后侧栏状态指示灯变绿。",
     "高"),
    ("HPC配比计算", "骨料用量计算",
     "输入Vg/ρg/ρs后mg和ms计算正确；参考表内容符合规范。",
     "高"),
    ("HPC配比计算", "胶凝材料计算",
     "输入各掺合料掺量和密度后，mc/m1/m2/m3/m4/mb结果合理；含气量参数对结果影响正确。",
     "高"),
    ("HPC配比计算", "水和外加剂与汇总",
     "mw=mb×W/B计算正确；外加剂用量按百分比计算正确；最终配比汇总表数据完整。",
     "高"),
    ("HPC配比计算", "保存与载入",
     "保存记录后可在项目详情和全部记录中查看；载入历史记录后所有参数正确填入计算器。",
     "高"),

    ("UHPC配比计算", "水胶比与推荐表",
     "点击推荐强度等级行自动带入W/B；外加剂未输入时自动补默认值。",
     "高"),
    ("UHPC配比计算", "砂胶比与钢纤维计算",
     "砂胶比输入后体积参数计算正确；钢纤维掺量与密度输入后质量计算正确。",
     "高"),
    ("UHPC配比计算", "胶凝材料比例(安德森模型)",
     "输入粒径(D10/D50/D90)和密度后体积比例→质量比例→每方用量全链路计算正确；微粉系数提示根据强度等级正确显示。",
     "高"),
    ("UHPC配比计算", "保存与载入",
     "同HPC：记录保存可查，载入参数完整正确。",
     "高"),

    ("HPC试配调整", "工作性实验",
     "调整Δmb/Δβs/Δα后试拌配合比和试拌用量实时更新(500ms防抖)；工作性评价参考正常显示。",
     "高"),
    ("HPC试配调整", "强度实验与回归",
     "试验组生成逻辑正确(+Δ/-Δ水胶比)；录入强度后回归分析结果合理(R²≥0.9时拟合良好)；图表正确渲染。",
     "高"),
    ("HPC试配调整", "配合比校正",
     "输入表观密度后校正系数δ计算正确；校正后实验室配合比各数值合理。",
     "中"),
    ("HPC试配调整", "试配保存与上下文",
     "试配记录正确保存并附着在主记录中；页面顶部项目名和记录名正确显示。",
     "中"),

    ("UHPC试配调整", "工作性实验",
     "调整S/B和α后试拌配合比重算正确。",
     "高"),
    ("UHPC试配调整", "强度实验与回归",
     "试验组生成(+Δ/-Δ w/b或硅灰掺量)正确；强度回归分析合理。",
     "高"),
    ("UHPC试配调整", "校正基准切换",
     "三种基准(试拌/推荐w/b/推荐硅灰)切换后校正结果不同且合理；表观密度校正正确。",
     "高"),
    ("UHPC试配调整", "试配保存",
     "最终实验室配合比正确保存为试配记录。",
     "中"),

    ("计算书", "计算书生成",
     "选择项目后正确列出记录；点击生成计算书弹出新窗口；HTML内容包含完整项目信息和计算参数键值表。",
     "中"),
    ("计算书", "打印/导出PDF",
     "计算书页面通过浏览器打印(Ctrl+P)可正常导出PDF。",
     "低"),

    ("全部配合记录", "记录列表与筛选",
     "全部记录表格12列数据完整；搜索(名称/项目)和类别筛选(HPC/UHPC)正常工作；分页正确。",
     "中"),
    ("全部配合记录", "载入功能",
     "点击载入按钮后正确跳转并填入数据。",
     "中"),

    ("回收站", "回收站列表",
     "已删除的项目和记录在回收站中可见；管理员视图显示全部用户数据；筛选和搜索功能正常。",
     "中"),
    ("回收站", "数据恢复",
     "恢复项目时关联记录一并恢复；恢复记录时若项目在回收站也一并恢复；弹窗提示影响范围正确。",
     "高"),
    ("回收站", "彻底删除",
     "彻底删除后数据不可恢复；弹窗警告正确；普通用户只能管理自己的数据。",
     "中"),

    ("个人中心", "个人资料修改",
     "邮箱和手机号可正常修改并保存；用户名只读不可修改。",
     "中"),
    ("个人中心", "密码修改",
     "输入正确原密码+两次一致新密码后修改成功；原密码错误时提示失败；修改成功后自动退出。",
     "高"),

    ("用户管理(管理员)", "用户列表",
     "表格正确展示所有用户信息(含角色和密码状态标签)；搜索过滤功能正常。",
     "中"),
    ("用户管理(管理员)", "新增用户",
     "创建用户后默认密码123456；用户名唯一性校验；邮箱格式校验。",
     "高"),
    ("用户管理(管理员)", "重置密码",
     "管理员可重置任意普通用户密码；两次密码一致性校验。",
     "高"),
    ("用户管理(管理员)", "删除用户",
     "可删除普通用户(二次确认)；管理员账号不可删除。",
     "中"),

    ("系统部署", "前端构建与部署",
     "执行npm run build:prod成功生成dist目录；dist内容部署到Nginx后可正常访问。",
     "中"),
    ("系统部署", "后端打包",
     "执行build.ps1/build.sh成功生成可执行文件main；运行main后API正常响应。",
     "中"),
    ("系统部署", "数据库初始化",
     "删除数据库文件后重启系统，自动重建全部表结构和默认管理员账号。",
     "高"),

    ("整体", "跨浏览器兼容性",
     "在Chrome/Edge/Firefox最新版本下页面布局正常、功能无异常。",
     "低"),
    ("整体", "响应式布局",
     "页面在1366×768及以上分辨率下无水平滚动条；关键操作区域不被遮挡。",
     "低"),
    ("整体", "操作流畅性",
     "页面切换、数据加载无明显卡顿(页面加载<3秒, 计算响应<2秒)。",
     "低"),
]


def generate_acceptance_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "验收清单"

    # 列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 26

    # ── 标题行 ──
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "水泥配比计算系统 —— 客户验收清单"
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = "版本: 2026-06-02 | 请逐项验收并在对应栏目签字/打勾确认"
    c.font = FONT_SMALL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # ── 客户信息行 ──
    info_row = 3
    info_labels = ["客户名称:", "验收日期:", "验收人:", "联系电话:"]
    info_positions = [(1, 2), (3, 4), (5, 6), (7, 8)]  # merge pairs
    for (c1, c2), label in zip(info_positions, info_labels):
        ws.merge_cells(f"{get_column_letter(c1)}{info_row}:{get_column_letter(c2)}{info_row}")
        cell = ws.cell(row=info_row, column=c1, value=label)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=DARK_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER
        for cc in range(c1, c2 + 1):
            ws.cell(row=info_row, column=cc).border = THIN_BORDER
    ws.row_dimensions[info_row].height = 30

    # ── 表头 ──
    header_row = 5
    headers = ["序号", "功能模块", "验收项", "验收标准与期望结果", "权重", "验收结果", "验收人签字", "备注"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 28

    # ── 数据行 ──
    row = 6
    last_module = ""
    seq = 0

    for module, item, criteria, weight in ACCEPTANCE_ITEMS:
        seq += 1
        is_section_first = (module != last_module)
        last_module = module

        values = [seq, module, item, criteria, weight, "", "", ""]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT if col_idx <= 4 else ALIGN_CENTER
            cell.border = THIN_BORDER

            if is_section_first and col_idx in (1, 2):
                cell.font = FONT_SECTION
                cell.fill = FILL_SECTION

        # 权重着色
        weight_cell = ws.cell(row=row, column=5)
        if weight == "高":
            weight_cell.font = Font(name="微软雅黑", size=10, bold=True, color=RED)
        elif weight == "中":
            weight_cell.font = Font(name="微软雅黑", size=10, bold=True, color="E65100")
        else:
            weight_cell.font = Font(name="微软雅黑", size=10, color="666666")

        # 验收结果列 — 浅绿色背景提示可填写
        for col_idx in (6, 7, 8):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = FILL_GREEN
            cell.font = Font(name="微软雅黑", size=10, italic=True, color="999999")

        # 奇偶行底色(仅前4列信息列)
        if seq % 2 == 0:
            for col_idx in range(1, 5):
                if not ws.cell(row=row, column=col_idx).fill or ws.cell(row=row, column=col_idx).fill.start_color.index == "00000000":
                    ws.cell(row=row, column=col_idx).fill = FILL_ROW_EVEN

        ws.row_dimensions[row].height = 42
        row += 1

    # ── 统计行 ──
    row += 1
    high_count = sum(1 for _, _, _, w in ACCEPTANCE_ITEMS if w == "高")
    mid_count = sum(1 for _, _, _, w in ACCEPTANCE_ITEMS if w == "中")
    low_count = sum(1 for _, _, _, w in ACCEPTANCE_ITEMS if w == "低")
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1,
                value=f"共 {len(ACCEPTANCE_ITEMS)} 项验收指标（高优先级 {high_count} 项 | 中优先级 {mid_count} 项 | 低优先级 {low_count} 项）")
    c.font = Font(name="微软雅黑", size=10, bold=True, color=DARK_BLUE)
    c.alignment = ALIGN_CENTER
    for cc in range(1, 9):
        ws.cell(row=row, column=cc).border = THIN_BORDER

    row += 2
    # 验收结论区
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1, value="验收结论:")
    c.font = Font(name="微软雅黑", size=12, bold=True, color=DARK_BLUE)
    c.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(f"D{row}:H{row}")
    options = ["□ 全部通过，准予交付", "□ 部分通过，需整改后复验", "□ 未通过，拒绝交付"]
    c = ws.cell(row=row, column=4, value="  |  ".join(options))
    c.font = Font(name="微软雅黑", size=11, color="333333")
    c.alignment = ALIGN_LEFT
    for cc in range(1, 9):
        ws.cell(row=row, column=cc).border = THIN_BORDER
    ws.row_dimensions[row].height = 36

    row += 2
    sign_labels = [
        ("客户代表签字:", 1, 2, "日期:", 3, 4),
        ("开发方代表签字:", 5, 6, "日期:", 7, 8),
    ]
    for sign_label, c1, c2, date_label, c3, c4 in sign_labels:
        ws.merge_cells(f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}")
        c = ws.cell(row=row, column=c1, value=sign_label)
        c.font = Font(name="微软雅黑", size=11, bold=True, color=DARK_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center")

        ws.merge_cells(f"{get_column_letter(c3)}{row}:{get_column_letter(c4)}{row}")
        c = ws.cell(row=row, column=c3, value=date_label)
        c.font = Font(name="微软雅黑", size=11, bold=True, color=DARK_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[row].height = 36
        row += 1

    # ── 冻结窗格 ──
    ws.freeze_panes = "A6"

    # ── 打印设置 ──
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    path = f"{OUT_DIR}/验收清单_2026-06-02.xlsx"
    wb.save(path)
    print(f"✅ 验收清单已生成: {path}")
    print(f"   共 {len(ACCEPTANCE_ITEMS)} 项验收指标 (高:{high_count} 中:{mid_count} 低:{low_count})")


if __name__ == "__main__":
    generate_feature_xlsx()
    generate_acceptance_xlsx()
    print("\n🎉 全部 Excel 文件已生成完毕！")
