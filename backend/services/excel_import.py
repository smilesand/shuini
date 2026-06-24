"""
Excel 导入与校验服务
===================
解析上传的 .xlsx 文件（横排分节版面，与 PDF 模板一致）。

版面采用截图模板中的五个分节：配合比信息、混凝土性能要求、原材料性能、
配合比关键参数、实验室配合比。导入后校验弹窗只展示关键合理性评价项。
"""

from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import load_workbook
from services.water_binder import calculate_water_binder_ratio, calculate_fcu0

# 校验容差
WB_TOLERANCE = 0.02      # 水胶比容差（图三要求）

# 导入合理性范围（缺省不阻断；有值则按范围评价）
SAND_RATIO_RANGE = (35.0, 50.0)          # 砂率 %
VG_FALLBACK_RANGE = (0.30, 0.40)         # 粗骨料体积用量 m³
SAND_BINDER_RATIO_RANGE = (0.80, 1.60)   # UHPC 砂胶比
STEEL_FIBER_RANGE = (0.0, 4.0)           # UHPC 钢纤维体积掺量 %


def _item(param: str, actual: Any, *, expected: Any = None, diff: Any = None,
          tolerance: Any = None, passed: bool = True) -> dict[str, Any]:
    """构造统一结构的校验项，确保前端所需字段齐全。"""
    return {
        "param": param,
        "expected": expected,
        "actual": actual,
        "diff": diff,
        "passed": passed,
        "tolerance": tolerance,
    }


def _num_from(raw: Any) -> float | None:
    """从可能带单位的文本/数值中解析首个数值（如 "20.0 mm" -> 20.0）。"""
    val = _parse_float(raw)
    if val is not None:
        return val
    if raw is None:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", str(raw))
    return float(match.group()) if match else None


def _fmt_num(value: float) -> str:
    """去除多余小数（5.0 -> 5）。"""
    return str(int(value)) if float(value).is_integer() else str(value)


def _range_check(param: str, raw: Any, lo: float, hi: float, unit: str,
                 items: list[dict[str, Any]], warnings: list[str]) -> bool:
    """范围校验：值缺省 -> 不生成校验项并视为合格；超范围 -> 生成不合格项 + 警告。

    返回 True 表示合格（或缺省），False 表示超出范围。
    """
    val = _num_from(raw)
    if val is None:
        return True
    range_text = f"{_fmt_num(lo)}~{_fmt_num(hi)} {unit}"
    ok = lo <= val <= hi
    items.append(_item(f"{param}范围", _round(val, 1), tolerance=range_text, passed=ok))
    if not ok:
        warnings.append(f"{param} {_fmt_num(val)} {unit} 超出合理范围（{range_text}）")
    return ok


def _range_item(param: str, actual: Any, lo: float, hi: float, unit: str = "") -> dict[str, Any]:
    val = _num_from(actual)
    if val is None:
        return _item(param, None, expected=None, diff=None, tolerance=f"{_fmt_num(lo)}~{_fmt_num(hi)}{unit}", passed=False)
    passed = lo <= val <= hi
    mid = round((lo + hi) / 2, 4)
    diff = abs(val - mid)
    return _item(
        param,
        _round(val, 4),
        expected=mid,
        diff=_round(diff, 4),
        tolerance=f"{_fmt_num(lo)}~{_fmt_num(hi)}{unit}",
        passed=passed,
    )


def _as_percent(value: Any) -> float | None:
    val = _num_from(value)
    if val is None:
        return None
    return val * 100 if 0 < val < 1 else val


def _vg_from_materials(data: dict[str, Any]) -> float | None:
    vg = _num_from(data.get("vg"))
    if vg is not None:
        return vg
    mg = _num_from(data.get("mg"))
    rhog = _num_from(data.get("rhog"))
    if mg is None or rhog is None or rhog <= 0:
        return None
    return mg / rhog


def _vg_range_from_requirement(data: dict[str, Any]) -> tuple[float, float]:
    slump = _num_from(data.get("req_slump"))
    spread = _num_from(data.get("req_spread"))
    if slump is not None and 180 <= slump <= 220:
        return 0.37, 0.40
    if spread is not None:
        if 500 <= spread <= 600:
            return 0.35, 0.38
        if 600 < spread <= 700:
            return 0.32, 0.36
        if 700 < spread <= 800:
            return 0.30, 0.32
    return VG_FALLBACK_RANGE


def _sand_range_from_aggregate(data: dict[str, Any]) -> tuple[float, float]:
    max_size = _num_from(data.get("max_aggregate_size"))
    fcuk = _num_from(data.get("fcuk") or data.get("strength_grade"))
    if fcuk is None or max_size is None:
        return SAND_RATIO_RANGE

    rows = [
        (80.0, [(45.0, 49.0), (44.0, 48.0), (43.0, 47.0)]),
        (90.0, [(43.0, 47.0), (42.0, 46.0), (41.0, 45.0)]),
        (100.0, [(41.0, 45.0), (40.0, 44.0), (39.0, 43.0)]),
    ]
    row_ranges = rows[-1][1]
    for strength_limit, ranges in rows:
        if fcuk <= strength_limit:
            row_ranges = ranges
            break

    if max_size <= 16:
        return row_ranges[0]
    if max_size <= 20:
        return row_ranges[1]
    return row_ranges[2]


# ── 横排表头映射：表头标签 → 内部字段名（可多键）──────────────────────────────
_HEADER_MAP: dict[str, list[str]] = {
    "强度等级/mpa": ["fcuk", "strength_grade"],
    "抗拉强度/mpa": ["tensile_strength"],
    "扩展度/mm": ["req_spread"],
    "坍落度/mm": ["req_slump"],
    "胶材28d强度/mpa": ["fb"],
    "粗骨料最大粒径/mm": ["max_aggregate_size"],
    "水胶比 w/b": ["wb", "water_binder_ratio"],
    "水胶比 w/b": ["wb", "water_binder_ratio"],
    "水胶比w/b": ["wb", "water_binder_ratio"],
    "水泥/%": ["bc", "bcp", "cement_pct"],
    "粉煤灰/%": ["b1p", "fly_ash_pct"],
    "矿粉/%": ["b2p", "slag_powder_pct"],
    "微珠/%": ["b3p", "micro_bead_pct"],
    "硅灰/%": ["b4p", "silica_fume_pct"],
    "胶材总量": ["mb", "total_binder"],
    "砂率/%": ["sand_ratio"],
    "砂率 βs/%": ["sand_ratio"],
    "粗骨料体积 /m³": ["vg"],
    "粗骨料体积 vg/m³": ["vg"],
    "外加剂/%": ["alpha", "admixture_ratio"],
    "外加剂掺量 α/%": ["alpha", "admixture_ratio"],
    "含气量/m3": ["air_content", "va"],
    "钢纤维/%": ["steel_fiber_volume_ratio"],
    "砂胶比": ["sand_binder_ratio"],
    "胶砂比": ["sand_binder_ratio"],
    "钢纤维体积掺量/%": ["steel_fiber_volume_ratio"],
}

# 同行 标签|值 映射（值在标签右侧单元格）
_INLINE_MAP: dict[str, str] = {
    "配比类别": "category",
    "记录名称": "record_name",
}


def _parse_float(value: Any) -> float | None:
    """尝试将值转为 float，失败返回 None。"""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        s = str(value).strip()
        if s == "":
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _store(result: dict[str, Any], keys: list[str], raw: Any) -> None:
    """按 keys 存值（数值优先；首次出现的有效值生效）。"""
    parsed = _parse_float(raw)
    if parsed is not None:
        for k in keys:
            result.setdefault(k, parsed)
        return
    if raw is not None and str(raw).strip():
        text = str(raw).strip()
        for k in keys:
            result.setdefault(k, text)


def _norm_label(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace(" ", "").lower()


def _is_section(row: list[Any], title: str) -> bool:
    return any(title in str(cell) for cell in row if cell is not None)


def _store_cell(result: dict[str, Any], key: str, raw: Any) -> None:
    _store(result, [key], raw)


def _find_value_row(rows: list[list[Any]], start: int, label: str) -> int | None:
    for idx in range(start, min(start + 8, len(rows))):
        if rows[idx] and str(rows[idx][0]).strip().startswith(label):
            return idx
    return None


def _parse_sections(rows: list[list[Any]], result: dict[str, Any]) -> None:
    category = str(result.get("category") or "").lower()

    for i, row in enumerate(rows):
        if _is_section(row, "配合比信息") and i + 2 < len(rows):
            headers = rows[i + 1]
            values = rows[i + 2]
            for c, header in enumerate(headers):
                label = _norm_label(header)
                raw = values[c] if c < len(values) else None
                if label == "配合比名称":
                    _store_cell(result, "record_name", raw)
                elif label == "配合比类型":
                    if raw is not None:
                        result["category"] = str(raw).strip().lower()
                        category = result["category"]
                elif label == "创建人":
                    _store_cell(result, "created_by", raw)
                elif label == "创建时间":
                    _store_cell(result, "created_at", raw)

        elif _is_section(row, "混凝土性能要求") and i + 2 < len(rows):
            headers = rows[i + 1]
            values = rows[i + 2]
            for c, header in enumerate(headers):
                label = _norm_label(header)
                raw = values[c] if c < len(values) else None
                if label == "强度等级/mpa":
                    _store(result, ["fcuk", "strength_grade"], _num_from(raw) if _num_from(raw) is not None else raw)
                elif label == "扩展度/mm":
                    _store_cell(result, "req_spread", raw)
                elif label == "坍落度/mm":
                    _store_cell(result, "req_slump", raw)
                elif label == "抗拉强度/mpa":
                    _store_cell(result, "tensile_strength", raw)

        elif _is_section(row, "原材料性能"):
            value_idx = i + 3 if i + 3 < len(rows) else None
            values = rows[value_idx] if value_idx is not None else []
            if category == "uhpc":
                headers = rows[i + 2] if i + 2 < len(rows) else []
                mapping = {
                    "水泥": "cement_density",
                    "粉煤灰": "fly_ash_density",
                    "微珠": "micro_bead_density",
                    "硅灰": "silica_fume_density",
                    "细骨料": "sand_density",
                    "体系最大粒径": "max_particle_size",
                    "粉煤灰峰值粒径": "fly_ash_peak_size",
                    "微珠峰值粒径": "micro_bead_peak_size",
                }
            else:
                headers = rows[i + 2] if i + 2 < len(rows) else []
                mapping = {
                    "水泥": "rhoc",
                    "粉煤灰": "rho1",
                    "矿粉": "rho2",
                    "微珠": "rho3",
                    "硅灰": "rho4",
                    "粗骨料": "rhog",
                    "细骨料": "rhos",
                }
                if values:
                    _store_cell(result, "fb", values[0] if len(values) > 0 else None)
                    _store_cell(result, "max_aggregate_size", values[9] if len(values) > 9 else None)

            for c, header in enumerate(headers):
                key = mapping.get(str(header).strip()) if header is not None else None
                if key and c < len(values):
                    _store_cell(result, key, values[c])

        elif _is_section(row, "配合比关键参数") and i + 2 < len(rows):
            headers = rows[i + 1]
            values = rows[i + 2]
            for c, header in enumerate(headers):
                label = _norm_label(header)
                raw = values[c] if c < len(values) else None
                keys = _HEADER_MAP.get(label)
                if keys:
                    _store(result, keys, raw)

        elif _is_section(row, "实验室配合比") and i + 1 < len(rows):
            headers = rows[i + 1]
            value_idx = _find_value_row(rows, i + 2, "每方用量")
            if value_idx is None:
                continue
            values = rows[value_idx]
            for c, header in enumerate(headers):
                label = _norm_label(header)
                raw = values[c] if c < len(values) else None
                if label == "水泥":
                    _store_cell(result, "mc", raw)
                elif label == "粉煤灰":
                    _store_cell(result, "m1", raw)
                elif label == "矿粉":
                    _store_cell(result, "m2", raw)
                elif label == "微珠":
                    _store_cell(result, "m3", raw)
                elif label == "硅灰":
                    _store_cell(result, "m4", raw)
                elif label == "粗骨料":
                    _store_cell(result, "mg", raw)
                elif label in ("细骨料", "细骨料(砂)"):
                    _store_cell(result, "ms", raw)
                elif label == "钢纤维":
                    _store_cell(result, "msf", raw)
                elif label == "水":
                    _store_cell(result, "mw", raw)
                elif label == "外加剂":
                    _store_cell(result, "ma", raw)
                elif label == "合计":
                    _store_cell(result, "total_mass", raw)


def parse_template_sheet(worksheet) -> dict[str, Any]:
    """
    解析横排分节模板 sheet。

    - 同行 标签|值：识别「配比类别」「记录名称」等内联键。
    - 表头行 + 数值行：表头标签命中映射时，读取下一行同列的数值。
    返回 { key: value } 字典。
    """
    rows: list[list[Any]] = [list(r) for r in worksheet.iter_rows(values_only=True)]
    result: dict[str, Any] = {}
    n = len(rows)

    for i, row in enumerate(rows):
        for c, cell in enumerate(row):
            if cell is None:
                continue
            label = str(cell).strip()
            if not label:
                continue
            norm = label.lower()

            # 内联 标签|值（同行右侧）
            inline_key = _INLINE_MAP.get(label)
            if inline_key is not None:
                value = row[c + 1] if c + 1 < len(row) else None
                if inline_key == "category":
                    if value is not None and str(value).strip():
                        result.setdefault("category", str(value).strip().lower())
                else:
                    _store(result, [inline_key], value)
                continue

            # 表头行 → 下一行同列取值（若紧邻行为空，回退取再下一行，兼容两级表头）
            keys = _HEADER_MAP.get(norm)
            if keys:
                value: Any = None
                for offset in (1, 2):
                    if i + offset >= n:
                        break
                    value_row = rows[i + offset]
                    candidate = value_row[c] if c < len(value_row) else None
                    if candidate is not None and str(candidate).strip():
                        value = candidate
                        break
                _store(result, keys, value)

    _parse_sections(rows, result)

    return result


def validate_hpc(data: dict[str, Any]) -> dict[str, Any]:
    """
    HPC 导入校验。
    图三要求弹窗只评价水胶比、砂率、粗骨料体积用量三个关键参数。
    """
    items: list[dict[str, Any]] = []
    warnings: list[str] = []

    fcuk = data.get("fcuk")
    fb = data.get("fb")
    wb = data.get("wb")
    sand_ratio = _as_percent(data.get("sand_ratio"))

    expected_wb = None
    wb_diff = None
    wb_passed = wb is not None
    if wb is None:
        warnings.append("缺少水胶比（W/B）")
    elif fcuk is not None and fb is not None:
        try:
            aa = data.get("aa", 0.33)
            ab = data.get("ab", 1.09)
            ac = data.get("ac", -49.54)
            expected_wb = round(calculate_water_binder_ratio(calculate_fcu0(fcuk), fb, aa, ab, ac), 4)
            wb_diff = round(abs(expected_wb - float(wb)), 4)
            wb_passed = wb_diff <= WB_TOLERANCE
        except (ValueError, TypeError):
            warnings.append("水胶比重算失败，请检查强度等级和胶材28d强度")
    else:
        warnings.append("缺少强度等级或胶材28d强度，水胶比仅按已填写值载入")
    items.append(_item("水胶比", _round(wb, 4), expected=expected_wb, diff=wb_diff, tolerance=WB_TOLERANCE, passed=wb_passed))

    sand_lo, sand_hi = _sand_range_from_aggregate(data)
    sand_item = _range_item("砂率", sand_ratio, sand_lo, sand_hi, "%")
    if sand_ratio is None:
        warnings.append("缺少砂率（βs）")
    items.append(sand_item)

    vg_value = _vg_from_materials(data)
    vg_lo, vg_hi = _vg_range_from_requirement(data)
    vg_item = _range_item("粗骨料体积用量/m³", vg_value, vg_lo, vg_hi)
    if vg_value is None:
        warnings.append("缺少粗骨料体积用量，且无法由粗骨料质量和密度推算")
    items.append(vg_item)

    valid = all(item.get("passed") for item in items)
    return {"valid": valid, "items": items, "warnings": warnings}


def validate_uhpc(data: dict[str, Any]) -> dict[str, Any]:
    """UHPC 导入校验。UHPC 模板无粗骨料，按水胶比和 UHPC 关键参数评价。"""
    items: list[dict[str, Any]] = []
    warnings: list[str] = []

    grade = data.get("strength_grade") or data.get("fcuk")
    wb = data.get("water_binder_ratio") or data.get("wb")
    sb = data.get("sand_binder_ratio")
    sf = data.get("steel_fiber_volume_ratio")
    alpha = data.get("admixture_ratio") or data.get("alpha")

    if grade is None:
        warnings.append("缺少强度等级")
    wb_passed = wb is not None
    if wb is None:
        warnings.append("缺少水胶比")
    items.append(_item("水胶比", _round(wb, 4), expected=None, diff=None, tolerance=WB_TOLERANCE, passed=wb_passed))

    if sb is None:
        warnings.append("缺少砂胶比")
    items.append(_range_item("砂胶比", sb, *SAND_BINDER_RATIO_RANGE))

    if sf is None:
        warnings.append("缺少钢纤维体积掺量")
    items.append(_range_item("钢纤维/%", sf, *STEEL_FIBER_RANGE, "%"))

    if alpha is None:
        warnings.append("缺少外加剂掺量")

    valid = grade is not None and all(item.get("passed") for item in items)
    return {"valid": valid, "items": items, "warnings": warnings}


def _round(value: Any, digits: int) -> Any:
    try:
        return round(float(value), digits)
    except (ValueError, TypeError):
        return value


def _detect_category(data: dict[str, Any]) -> str:
    category = data.get("category", "hpc")
    if isinstance(category, str):
        category = category.lower().strip()
    if category not in ("hpc", "uhpc"):
        category = "hpc"
    return category


def parse_and_validate_excel(file_bytes: bytes, filename: str = "") -> dict[str, Any]:
    """
    解析上传的 Excel 文件并执行校验（单条记录横排模板格式）。

    返回: {
        "category": "hpc" | "uhpc",
        "data": {...},
        "validation": { "valid": bool, "items": [...], "warnings": [...] },
    }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    target_sheet = None
    for name in ["配比导入模板", "Sheet1", sheet_names[0]]:
        if name in sheet_names:
            target_sheet = name
            break
    if target_sheet is None:
        target_sheet = sheet_names[0]

    ws = wb[target_sheet]
    data = parse_template_sheet(ws)
    wb.close()

    category = _detect_category(data)
    data["category"] = category
    validation = validate_uhpc(data) if category == "uhpc" else validate_hpc(data)

    return {
        "category": category,
        "data": data,
        "validation": validation,
    }


# ── 项目导入（多 sheet）──────────────────────────────────────────────────────

PROJECT_INFO_LABEL_MAP = {
    "项目编号": "project_code",
    "项目名称": "project_name",
    "技术要求": "requirements",
    "创建人": "created_by",
    "创建时间": "created_at",
    "导出时间": "exported_at",
    "记录数量": "record_count",
}


def parse_project_info_sheet(worksheet) -> dict[str, Any]:
    """解析「项目信息」sheet（简单 key-value 两列布局）。"""
    result: dict[str, Any] = {}
    for row in worksheet.iter_rows(min_row=1, max_col=2, values_only=True):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        value = row[1] if len(row) > 1 else None
        key = PROJECT_INFO_LABEL_MAP.get(label)
        if key:
            result[key] = str(value).strip() if value is not None else ""
    return result


def parse_project_import_excel(file_bytes: bytes) -> dict[str, Any]:
    """
    解析项目导入 Excel（多 sheet 格式，与项目导出对应）。

    Sheet 1 「项目信息」— 项目元数据
    后续 sheet — 各条记录（横排模板格式）
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    if len(sheet_names) < 1:
        wb.close()
        raise ValueError("Excel 文件中没有 sheet")

    project_info = parse_project_info_sheet(wb[sheet_names[0]])

    records: list[dict[str, Any]] = []
    all_valid = True

    for sheet_name in sheet_names[1:]:
        ws = wb[sheet_name]
        data = parse_template_sheet(ws)
        if not data:
            continue

        category = _detect_category(data)
        data["category"] = category
        validation = validate_uhpc(data) if category == "uhpc" else validate_hpc(data)

        name = data.get("record_name") or sheet_name
        if not validation.get("valid"):
            all_valid = False

        records.append({
            "category": category,
            "name": str(name),
            "data": data,
            "validation": validation,
        })

    wb.close()

    return {
        "project": project_info,
        "records": records,
        "all_valid": all_valid,
    }
