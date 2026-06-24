"""
Excel 模板服务
=============
生成配比导入模板。模板版面与 PDF 报告一致：分节横排（配合比信息 / 混凝土性能要求 / 原材料性能 / 配合比关键参数 / 实验室配合比）。
使用 openpyxl 生成 .xlsx 文件。
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 样式常量 ──────────────────────────────────────────────────────────────────
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
VALUE_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
VALUE_FONT = Font(name="微软雅黑", size=10)
LABEL_FONT = Font(name="微软雅黑", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── 工具函数 ──────────────────────────────────────────────────────────────────
def _flatten(obj: Any, out: dict[str, Any] | None = None) -> dict[str, Any]:
    """递归扁平化嵌套 dict，叶子值写入同一层字典（与前端报告一致）。"""
    if out is None:
        out = {}
    if not isinstance(obj, dict):
        return out
    for k, v in obj.items():
        if isinstance(v, dict):
            _flatten(v, out)
        elif not isinstance(v, list):
            out.setdefault(k, v)
    return out


def _first(flat: dict[str, Any], *keys: str) -> Any:
    """返回第一个存在且非 None 的键值。"""
    for k in keys:
        if k in flat and flat[k] is not None:
            return flat[k]
    return None


def _num(value: Any, digits: int = 2) -> Any:
    """转为四舍五入后的数值；无法转换时返回原值（字符串）或 None。"""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return round(float(value), digits)
    try:
        return round(float(str(value).strip()), digits)
    except (ValueError, TypeError):
        return value


def _pct(value: Any, digits: int = 1) -> Any:
    """归一化百分比：小于 1 的小数视为比例并 ×100。"""
    n = _num(value, 6)
    if isinstance(n, (int, float)) and not isinstance(n, bool) and n < 1:
        return round(n * 100, digits)
    return _num(value, digits)


def _style(cell, font=None, fill=None, alignment=None, border=THIN_BORDER):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def _section_bar(ws, row: int, ncol: int, title: str) -> int:
    """写入区块标题栏（跨 ncol 列合并）。返回下一行号。"""
    cell = ws.cell(row=row, column=1, value=title)
    _style(cell, font=SECTION_FONT, fill=SECTION_FILL, alignment=CENTER_ALIGN)
    if ncol > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    for c in range(1, ncol + 1):
        _style(ws.cell(row=row, column=c))
    return row + 1


def _write_row(ws, row: int, values: list[Any], header: bool = False,
               input_cells: set[int] | None = None) -> None:
    """写入一行数据。header=True 使用表头样式；input_cells 中的列使用浅黄底（填写区）。"""
    input_cells = input_cells or set()
    for i, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=val)
        if header:
            _style(cell, font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
        elif i in input_cells:
            _style(cell, font=VALUE_FONT, fill=INPUT_FILL, alignment=CENTER_ALIGN)
        else:
            _style(cell, font=VALUE_FONT, alignment=CENTER_ALIGN)


def _write_filled_row(ws, row: int, values: list[Any], fill: PatternFill) -> None:
    for i, val in enumerate(values, start=1):
        _style(ws.cell(row=row, column=i, value=val), font=VALUE_FONT, fill=fill, alignment=CENTER_ALIGN)


# ── 区块构建 ──────────────────────────────────────────────────────────────────
def _build_hpc_sheet(ws, name: str, flat: dict[str, Any]) -> None:
    ncol = 11
    for i in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions["A"].width = 16

    cell = ws.cell(row=1, column=1, value="混凝土配合比记录 · 导入模板")
    _style(cell, font=TITLE_FONT, alignment=CENTER_ALIGN, border=THIN_BORDER)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    row = 2

    fcuk = _num(_first(flat, "fcuk", "strengthGrade", "strength_grade"), 1)
    created_by = _first(flat, "created_by") or "admin"
    created_at = _first(flat, "created_at") or datetime.now().strftime("%Y.%m.%d")

    row = _section_bar(ws, row, ncol, "一、 配合比信息")
    _write_row(ws, row, ["配合比名称", "配合比类型", "创建人", "创建时间"], header=True)
    row += 1
    _write_filled_row(ws, row, [name, "HPC", created_by, created_at], VALUE_FILL)
    row += 2

    row = _section_bar(ws, row, ncol, "二、 混凝土性能要求")
    _write_row(ws, row, ["强度等级/MPa", "扩展度/mm", "坍落度/mm", "其他"], header=True)
    row += 1
    _write_filled_row(ws, row, [fcuk, _num(_first(flat, "req_spread", "reqSpread"), 0), _num(_first(flat, "req_slump", "reqSlump"), 0), ""], VALUE_FILL)
    row += 2

    row = _section_bar(ws, row, ncol, "三、 原材料性能")
    _style(ws.cell(row=row, column=1, value="胶材28d强度/MPa"), font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    _style(ws.cell(row=row, column=2, value="表观密度/kg/m³"), font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    for c in range(2, 9):
        _style(ws.cell(row=row, column=c), font=HEADER_FONT, fill=HEADER_FILL)
    _style(ws.cell(row=row, column=9, value="粗骨料最大粒径/mm"), font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    ws.merge_cells(start_row=row, start_column=9, end_row=row + 1, end_column=9)
    row += 1
    for i, sub in enumerate(["", "水泥", "粉煤灰", "矿粉", "微珠", "硅灰", "粗骨料", "细骨料", ""], start=1):
        if i in (1, 9):
            _style(ws.cell(row=row, column=i), font=HEADER_FONT, fill=HEADER_FILL)
            continue
        _style(ws.cell(row=row, column=i, value=sub), font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    row += 1
    _write_filled_row(ws, row, [
        _num(_first(flat, "fb", "binderStrength28d"), 1),
        _num(flat.get("rhoc"), 0), _num(flat.get("rho1"), 0), _num(flat.get("rho2"), 0),
        _num(flat.get("rho3"), 0), _num(flat.get("rho4"), 0), _num(flat.get("rhog"), 0),
        _num(flat.get("rhos"), 0),
        _first(flat, "max_aggregate_size", "maxAggregateSize"),
    ], VALUE_FILL)
    row += 2

    row = _section_bar(ws, row, ncol, "四、 配合比关键参数")
    _write_row(ws, row, ["水胶比 W/B", "水泥/%", "粉煤灰/%", "矿粉/%", "微珠/%", "硅灰/%", "胶材总量", "砂率/%", "粗骨料体积 /m³", "外加剂/%", "含气量/m3"], header=True)
    row += 1
    _write_filled_row(ws, row, [
        _num(_first(flat, "wb", "water_binder_ratio"), 3),
        _pct(_first(flat, "bc", "bcp", "cement_pct")),
        _pct(_first(flat, "b1p", "fly_ash_pct")),
        _pct(_first(flat, "b2p", "slag_powder_pct")),
        _pct(_first(flat, "b3p", "micro_bead_pct")),
        _pct(_first(flat, "b4p", "silica_fume_pct")),
        _num(_first(flat, "mb", "total_binder"), 0),
        _pct(_first(flat, "sand_ratio", "sandRatio")),
        _num(flat.get("vg"), 2),
        _pct(_first(flat, "alpha", "admixture_ratio"), 2),
        _num(_first(flat, "air_content", "va"), 2),
    ], INPUT_FILL)
    row += 2

    row = _section_bar(ws, row, ncol, "五、 实验室配合比")
    _write_row(ws, row, ["用量", "水泥", "粉煤灰", "矿粉", "微珠", "硅灰", "粗骨料", "细骨料", "水", "外加剂", "合计"], header=True)
    row += 1
    keys = ["mc", "m1", "m2", "m3", "m4", "mg", "ms", "mw", "ma", "total_mass"]
    per_m3 = [_num(_first(flat, k, "totalMass") if k == "total_mass" else flat.get(k), 2) for k in keys]
    _write_filled_row(ws, row, ["每方用量(kg/m³)", *per_m3], VALUE_FILL)
    row += 1
    v_batch = _num(_first(flat, "batchVolume", "batch_volume"), 0) or 20
    scale = v_batch / 1000.0
    batch = [round(x * scale, 3) if isinstance(x, (int, float)) and not isinstance(x, bool) else None for x in per_m3]
    _write_filled_row(ws, row, [f"试拌用量(kg/{int(v_batch)}L)", *batch], INPUT_FILL)


def _build_uhpc_sheet(ws, name: str, flat: dict[str, Any]) -> None:
    ncol = 10
    for i in range(1, ncol + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions["A"].width = 16

    cell = ws.cell(row=1, column=1, value="混凝土配合比记录 · 导入模板")
    _style(cell, font=TITLE_FONT, alignment=CENTER_ALIGN, border=THIN_BORDER)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    row = 2

    grade = _num(_first(flat, "strengthGrade", "strength_grade", "fcuk"), 0)
    created_by = _first(flat, "created_by") or "admin"
    created_at = _first(flat, "created_at") or datetime.now().strftime("%Y.%m.%d")

    row = _section_bar(ws, row, ncol, "一、 配合比信息")
    _write_row(ws, row, ["配合比名称", "配合比类型", "创建人", "创建时间"], header=True)
    row += 1
    _write_filled_row(ws, row, [name, "UHPC", created_by, created_at], VALUE_FILL)
    row += 2

    row = _section_bar(ws, row, ncol, "二、 混凝土性能要求")
    _write_row(ws, row, ["强度等级/MPa", "扩展度/mm", "抗拉强度/MPa", "其他"], header=True)
    row += 1
    _write_filled_row(ws, row, [grade, _num(_first(flat, "req_spread", "reqSpread"), 0), _num(_first(flat, "tensile_strength"), 1), ""], VALUE_FILL)
    row += 2

    row = _section_bar(ws, row, ncol, "三、 原材料性能")
    _style(ws.cell(row=row, column=1, value="表观密度/kg/m³"), font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    for c in range(1, 6):
        _style(ws.cell(row=row, column=c), font=HEADER_FONT, fill=HEADER_FILL)
    _style(ws.cell(row=row, column=6, value="粒径/μm"), font=HEADER_FONT, fill=HEADER_FILL, alignment=CENTER_ALIGN)
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
    for c in range(6, 9):
        _style(ws.cell(row=row, column=c), font=HEADER_FONT, fill=HEADER_FILL)
    row += 1
    _write_row(ws, row, ["水泥", "粉煤灰", "微珠", "硅灰", "细骨料", "体系最大粒径", "粉煤灰峰值粒径", "微珠峰值粒径"], header=True)
    row += 1
    _write_filled_row(ws, row, [
        _num(_first(flat, "cement_density", "rhoc"), 0),
        _num(_first(flat, "fly_ash_density", "rho1"), 0),
        _num(_first(flat, "micro_bead_density", "rho3"), 0),
        _num(_first(flat, "silica_fume_density", "rho4"), 0),
        _num(_first(flat, "sand_density", "rhos"), 0),
        _num(_first(flat, "max_particle_size"), 0),
        _num(_first(flat, "fly_ash_peak_size"), 0),
        _num(_first(flat, "micro_bead_peak_size"), 0),
    ], VALUE_FILL)
    row += 2

    row = _section_bar(ws, row, ncol, "四、 配合比关键参数")
    _write_row(ws, row, ["水胶比 W/B", "水泥/%", "粉煤灰/%", "微珠/%", "硅灰/%", "胶材总量", "钢纤维/%", "砂胶比", "外加剂/%", ""], header=True)
    row += 1
    _write_filled_row(ws, row, [
        _num(_first(flat, "waterBinderRatio", "water_binder_ratio", "wb"), 3),
        _pct(_first(flat, "bc", "bcp", "cement_pct")),
        _pct(_first(flat, "b1p", "fly_ash_pct")),
        _pct(_first(flat, "b3p", "micro_bead_pct")),
        _pct(_first(flat, "b4p", "silica_fume_pct")),
        _num(_first(flat, "mb", "total_binder"), 0),
        _pct(_first(flat, "steelFiberVolumeRatio", "steel_fiber_volume_ratio"), 2),
        _num(_first(flat, "sandBinderRatio", "sand_binder_ratio", "sand_ratio"), 2),
        _pct(_first(flat, "admixtureRatio", "admixture_ratio", "alpha"), 2),
        "",
    ], INPUT_FILL)
    row += 2

    row = _section_bar(ws, row, ncol, "五、 实验室配合比")
    _write_row(ws, row, ["用量", "水泥", "粉煤灰", "微珠", "硅灰", "细骨料", "钢纤维", "水", "外加剂", "合计"], header=True)
    row += 1
    keys = ["mc", "m1", "m3", "m4", "ms", "msf", "mw", "ma", "total"]
    per_m3 = [_num(_first(flat, "total", "totalMass", "total_mass") if k == "total" else flat.get(k), 2) for k in keys]
    _write_filled_row(ws, row, ["每方用量(kg/m³)", *per_m3], VALUE_FILL)
    row += 1
    v_batch = _num(_first(flat, "batchVolume", "batch_volume"), 0) or 20
    scale = v_batch / 1000.0
    batch = [round(x * scale, 3) if isinstance(x, (int, float)) and not isinstance(x, bool) else None for x in per_m3]
    _write_filled_row(ws, row, [f"试拌用量(kg/{int(v_batch)}L)", *batch], INPUT_FILL)


def _build_record_sheet(ws, record: dict[str, Any]) -> None:
    category = (record.get("category") or "hpc").lower()
    name = record.get("name", "未命名")
    flat = _flatten(record.get("record_data", {}) or {})
    if category.startswith("uhpc"):
        _build_uhpc_sheet(ws, name, flat)
    else:
        _build_hpc_sheet(ws, name, flat)


# ── 公开接口 ──────────────────────────────────────────────────────────────────
def generate_template_bytes(category: str) -> bytes:
    """生成空白导入模板（横排版面，仅需填写关键参数填写区）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "配比导入模板"
    ws.sheet_properties.tabColor = "4472C4"
    if category == "uhpc":
        _build_uhpc_sheet(ws, "导入模板", {})
    else:
        _build_hpc_sheet(ws, "导入模板", {})

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_record_export_bytes(record: dict[str, Any], project_name: str = "") -> bytes:
    """生成单条记录导出 Excel（横排版面，与 PDF 一致）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(record.get("name", "记录"))
    ws.sheet_properties.tabColor = "4472C4"
    _build_record_sheet(ws, record)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_project_export_bytes(project: dict[str, Any], records: list[dict[str, Any]]) -> bytes:
    """生成项目导出 Excel（多 sheet：项目信息 + 每条记录一个横排 sheet）。"""
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "项目信息"
    ws_info.sheet_properties.tabColor = "4472C4"
    ws_info.column_dimensions["A"].width = 20
    ws_info.column_dimensions["B"].width = 40

    info_data = [
        ("项目编号", project.get("project_code", "")),
        ("项目名称", project.get("project_name", "")),
        ("技术要求", project.get("requirements", "")),
        ("创建人", project.get("created_by", "")),
        ("创建时间", project.get("created_at", "")),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("记录数量", str(len(records))),
    ]
    for i, (label, value) in enumerate(info_data, start=1):
        _style(ws_info.cell(row=i, column=1, value=label), font=LABEL_FONT, alignment=LEFT_ALIGN)
        _style(ws_info.cell(row=i, column=2, value=value), font=VALUE_FONT, alignment=LEFT_ALIGN)

    for record in records:
        base_name = _safe_sheet_name(record.get("name", f"记录{record.get('id', '')}"))
        sheet_name = base_name
        counter = 1
        while sheet_name in wb.sheetnames:
            counter += 1
            sheet_name = f"{base_name[:28]}_{counter}"
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = "4472C4"
        _build_record_sheet(ws, record)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """生成合法的 Excel sheet 名称。"""
    forbidden = r'[]:*?/\\'
    safe = "".join(c for c in name if c not in forbidden)
    safe = safe.strip("'")
    if not safe:
        safe = "Sheet"
    return safe[:max_len]


def _export_filename(prefix: str) -> str:
    """生成带时间戳的导出文件名。"""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.xlsx"
