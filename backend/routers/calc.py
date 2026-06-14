import os
import tempfile

from fastapi import APIRouter, HTTPException, UploadFile, File
from models.schemas import (
    WaterBinderRequest, WaterBinderResponse,
    FitCoefficientsRequest, FitCoefficientsResponse,
    AggregateRequest, AggregateResponse,
    BinderRequest, BinderResponse,
    WaterAdmixtureRequest, WaterAdmixtureResponse,
    UhpcMixRequest, UhpcMixResponse,
    AdaptRequest, AdaptResponse,
    HpcTrialRequest, HpcTrialResponse,
    UhpcTrialRequest, UhpcTrialResponse,
    ok,
)
from services.calculations import (
    calc_water_binder,
    fit_regression_coefficients,
    calc_aggregate,
    calc_binder,
    calc_water_admixture,
    calc_uhpc_mix,
    calc_hpc_trial,
    calc_uhpc_trial,
)
from services.adapt import calc_adapt

MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

router = APIRouter()


@router.post("/water-binder", summary="计算水胶比")
def api_water_binder(req: WaterBinderRequest):
    try:
        result = calc_water_binder(req.fcuk, req.fb, req.aa, req.ab, req.ac)
        return ok(WaterBinderResponse(**result).model_dump())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="水胶比计算失败")


@router.post("/fit-coefficients", summary="拟合回归系数")
def api_fit_coefficients(req: FitCoefficientsRequest):
    try:
        result = fit_regression_coefficients(req.csv_text)
        return ok(FitCoefficientsResponse(**result).model_dump())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="回归系数拟合失败")


@router.post("/aggregate", summary="计算骨料用量")
def api_aggregate(req: AggregateRequest):
    try:
        result = calc_aggregate(req.vg, req.rhog, req.beta_s, req.rhos)
        return ok(AggregateResponse(**result).model_dump())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="骨料计算失败")


@router.post("/binder", summary="计算胶凝材料")
def api_binder(req: BinderRequest):
    try:
        result = calc_binder(
            req.b1p, req.rho1, req.b2p, req.rho2,
            req.b3p, req.rho3, req.b4p, req.rho4,
            req.rhoc, req.va,
            req.mg, req.ms, req.rhog, req.rhos,
            req.wb,
        )
        return ok(BinderResponse(**result).model_dump())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="胶凝材料计算失败")


@router.post("/water-admixture", summary="计算用水量与外加剂")
def api_water_admixture(req: WaterAdmixtureRequest):
    try:
        result = calc_water_admixture(req.mb, req.wb, req.alpha)
        return ok(WaterAdmixtureResponse(**result).model_dump())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="用水量计算失败")


@router.post("/uhpc-mix", summary="计算超高性能混凝土配合比")
def api_uhpc_mix(req: UhpcMixRequest):
    try:
        result = calc_uhpc_mix(**req.model_dump())
        return ok(UhpcMixResponse(**result).model_dump())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="超高性能混凝土配合比计算失败")


@router.post("/upload-fit", summary="上传 Excel/CSV 拟合回归系数")
async def api_upload_fit(file: UploadFile = File(...)):
    """上传 .xlsx 或 .csv 文件，解析后拟合回归系数，解析完后删除文件"""
    # 校验文件名
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"不支持的文件类型 {ext}，仅支持 {', '.join(ALLOWED_EXTENSIONS)}")

    # 读取并校验文件大小
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail=f"文件过大（最大 {MAX_UPLOAD_BYTES // 1024 // 1024} MB）")

    # 保存临时文件
    suffix = ext if ext in (".csv", ".xlsx", ".xls") else ".csv"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        if ext in (".xlsx", ".xls"):
            csv_text = _parse_xlsx(tmp_path)
        else:
            with open(tmp_path, encoding="utf-8-sig") as f:
                csv_text = f.read()
        if not csv_text.strip():
            raise HTTPException(status_code=422, detail="文件内容为空")
        result = fit_regression_coefficients(csv_text)
        return ok(FitCoefficientsResponse(**result).model_dump())
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="文件处理失败")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


def _parse_xlsx(path: str) -> str:
    """将 Excel 文件转换为 CSV 文本"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise ValueError("Excel 文件中没有活动工作表")
    lines = []
    for row in ws.iter_rows(min_row=2 if ws.max_row > 1 and _is_header(ws[1]) else 1):
        vals = [str(c.value or "").strip() for c in row[:3]]
        if all(vals):
            lines.append(",".join(vals))
    wb.close()
    if len(lines) < 3:
        raise ValueError(f"Excel 有效数据行不足（{len(lines)} 行），至少需要 3 行")
    return "\n".join(lines)


def _is_header(row) -> bool:
    """判断第一行是否为表头"""
    try:
        float(str(row[0].value or "").strip())
        return False
    except ValueError:
        return True


@router.post("/adapt", summary="适配调整计算（试拌配合比）")
def api_adapt(req: AdaptRequest):
    try:
        result = calc_adapt(
            mb_adj=req.mb_adj,
            beta_s_adj=req.beta_s_adj,
            alpha_adj=req.alpha_adj / 100.0,
            wb=req.wb,
            bc=req.bc,
            b1=req.b1,
            b2=req.b2,
            b3=req.b3,
            b4=req.b4,
            mg=req.mg,
        )
        return ok(AdaptResponse(**result).model_dump())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="适配调整计算失败")


@router.post("/hpc-trial", summary="计算高性能试配")
def api_hpc_trial(req: HpcTrialRequest):
    """
    统一计算高性能试配三步结果。

    该接口负责把工作性实验、强度实验、配合比校正与确认共用的业务公式集中到服务端，
    让前端改为通过节流防抖请求后端结果，而不是在浏览器中重复维护公式。
    """
    try:
        result = calc_hpc_trial(**req.model_dump())
        return ok(HpcTrialResponse(**result).model_dump())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="高性能试配计算失败")


@router.post("/uhpc-trial", summary="计算超高性能试配")
def api_uhpc_trial(req: UhpcTrialRequest):
    """统一计算超高性能试配三步结果。"""
    try:
        result = calc_uhpc_trial(**req.model_dump())
        return ok(UhpcTrialResponse(**result).model_dump())
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=500, detail="超高性能试配计算失败")
